# -*- coding: utf-8 -*-
"""
Excel 파일 전체 데이터 추출 스크립트
시트별로 분리된 파일로 저장하여 빠짐없이 추출
"""
import sys
import io
import os
import json

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import openpyxl
import xlrd
import warnings
warnings.filterwarnings('ignore')

BASE_PATH = r'C:\Users\rando\repos\reAA'
OUTPUT_DIR = os.path.join(BASE_PATH, 'extracted_data')

def ensure_dir(path):
    if not os.path.exists(path):
        os.makedirs(path)

def extract_xlsm_sheet_full(wb, sheet_name, output_path):
    """xlsm 시트의 모든 데이터 추출"""
    ws = wb[sheet_name]

    sheet_data = {
        'sheet_name': sheet_name,
        'dimensions': ws.dimensions,
        'row_count': ws.max_row,
        'col_count': ws.max_column,
        'all_cells': [],
        'formulas': [],
        'input_cells': [],
        'output_cells': [],
        'merged_cells': [str(mc) for mc in ws.merged_cells.ranges]
    }

    # 모든 셀 데이터 추출
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if cell.value is not None:
                cell_info = {
                    'coord': cell.coordinate,
                    'row': row_idx,
                    'col': col_idx,
                    'value': str(cell.value)[:500] if cell.value else None,
                    'data_type': cell.data_type,
                    'is_formula': str(cell.value).startswith('=') if cell.value else False
                }

                sheet_data['all_cells'].append(cell_info)

                if cell_info['is_formula']:
                    sheet_data['formulas'].append({
                        'coord': cell.coordinate,
                        'formula': str(cell.value)
                    })
                    sheet_data['output_cells'].append(cell.coordinate)
                else:
                    sheet_data['input_cells'].append(cell.coordinate)

    # 저장
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(sheet_data, f, ensure_ascii=False, indent=2)

    return len(sheet_data['all_cells']), len(sheet_data['formulas'])

def extract_xls_sheet_full(wb, sheet_name, output_path):
    """xls 시트의 모든 데이터 추출"""
    ws = wb.sheet_by_name(sheet_name)

    sheet_data = {
        'sheet_name': sheet_name,
        'dimensions': f'A1:{xlrd.colname(ws.ncols-1) if ws.ncols > 0 else "A"}{ws.nrows}',
        'row_count': ws.nrows,
        'col_count': ws.ncols,
        'all_cells': [],
        'input_cells': []
    }

    # 모든 셀 데이터 추출
    for row_idx in range(ws.nrows):
        for col_idx in range(ws.ncols):
            cell = ws.cell(row_idx, col_idx)

            if cell.value not in (None, ''):
                cell_info = {
                    'coord': f'{xlrd.colname(col_idx)}{row_idx+1}',
                    'row': row_idx + 1,
                    'col': col_idx + 1,
                    'value': str(cell.value)[:500],
                    'cell_type': cell.ctype,
                    'cell_type_name': ['EMPTY', 'TEXT', 'NUMBER', 'DATE', 'BOOLEAN', 'ERROR', 'BLANK'][cell.ctype]
                }

                sheet_data['all_cells'].append(cell_info)
                sheet_data['input_cells'].append(cell_info['coord'])

    # 저장
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(sheet_data, f, ensure_ascii=False, indent=2)

    return len(sheet_data['all_cells']), 0

def main():
    ensure_dir(OUTPUT_DIR)

    # 요약 정보
    summary = {
        'files': [],
        'total_sheets': 0,
        'total_cells': 0,
        'total_formulas': 0
    }

    # 1. xlsm 파일 처리
    xlsm_file = os.path.join(BASE_PATH, '경주대종천하류보20250501_unlocked.xlsm')
    if os.path.exists(xlsm_file):
        print(f"Processing: {os.path.basename(xlsm_file)}")
        wb = openpyxl.load_workbook(xlsm_file, data_only=False)

        file_dir = os.path.join(OUTPUT_DIR, 'xlsm_unlocked')
        ensure_dir(file_dir)

        file_info = {
            'filename': os.path.basename(xlsm_file),
            'type': 'xlsm',
            'sheets': []
        }

        for sheet_name in wb.sheetnames:
            safe_name = sheet_name.replace('/', '_').replace('\\', '_')
            output_path = os.path.join(file_dir, f'{safe_name}.json')

            cell_count, formula_count = extract_xlsm_sheet_full(wb, sheet_name, output_path)

            file_info['sheets'].append({
                'name': sheet_name,
                'output_file': output_path,
                'cell_count': cell_count,
                'formula_count': formula_count
            })

            summary['total_cells'] += cell_count
            summary['total_formulas'] += formula_count
            print(f"  - {sheet_name}: {cell_count} cells, {formula_count} formulas")

        summary['files'].append(file_info)
        summary['total_sheets'] += len(wb.sheetnames)
        wb.close()

    # 2. xls 파일들 처리
    xls_files = [
        os.path.join(BASE_PATH, '경주_대종천_하류보.xls'),
        os.path.join(BASE_PATH, '경주대종천하류보20250501.xls')
    ]

    for xls_file in xls_files:
        if os.path.exists(xls_file):
            print(f"\nProcessing: {os.path.basename(xls_file)}")
            wb = xlrd.open_workbook(xls_file)

            safe_filename = os.path.basename(xls_file).replace('.xls', '').replace(' ', '_')
            file_dir = os.path.join(OUTPUT_DIR, f'xls_{safe_filename}')
            ensure_dir(file_dir)

            file_info = {
                'filename': os.path.basename(xls_file),
                'type': 'xls',
                'sheets': []
            }

            for sheet_name in wb.sheet_names():
                safe_name = sheet_name.replace('/', '_').replace('\\', '_')
                output_path = os.path.join(file_dir, f'{safe_name}.json')

                cell_count, formula_count = extract_xls_sheet_full(wb, sheet_name, output_path)

                file_info['sheets'].append({
                    'name': sheet_name,
                    'output_file': output_path,
                    'cell_count': cell_count,
                    'formula_count': formula_count
                })

                summary['total_cells'] += cell_count
                print(f"  - {sheet_name}: {cell_count} cells")

            summary['files'].append(file_info)
            summary['total_sheets'] += len(wb.sheet_names())

    # 요약 저장
    summary_path = os.path.join(OUTPUT_DIR, 'extraction_summary.json')
    with open(summary_path, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(f"\n{'='*60}")
    print(f"Extraction complete!")
    print(f"Total files: {len(summary['files'])}")
    print(f"Total sheets: {summary['total_sheets']}")
    print(f"Total cells with data: {summary['total_cells']}")
    print(f"Total formulas: {summary['total_formulas']}")
    print(f"Output directory: {OUTPUT_DIR}")
    print(f"Summary file: {summary_path}")

if __name__ == '__main__':
    main()
