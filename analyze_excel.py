# -*- coding: utf-8 -*-
"""
Excel 파일 구조 분석 스크립트
모든 출력은 UTF-8 인코딩된 파일로 저장
"""
import sys
import io
import os
import json

# 표준 출력 인코딩 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import openpyxl
import warnings
warnings.filterwarnings('ignore')

def analyze_xlsm_file(file_path):
    """xlsm/xlsx 파일 상세 분석"""
    result = {}

    wb = openpyxl.load_workbook(file_path, data_only=False)

    result['filename'] = os.path.basename(file_path)
    result['filepath'] = file_path
    result['sheet_count'] = len(wb.sheetnames)
    result['sheet_names'] = wb.sheetnames
    result['sheets'] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            'dimensions': ws.dimensions,
            'min_row': ws.min_row,
            'max_row': ws.max_row,
            'min_col': ws.min_column,
            'max_col': ws.max_column,
            'headers': [],
            'sample_data': [],
            'formulas': [],
            'input_cells': [],  # 사용자 입력으로 추정되는 셀
            'output_cells': []  # 수식 결과로 추정되는 셀
        }

        # 첫 5행 데이터 샘플링 (헤더 및 데이터 파악)
        for row_idx, row in enumerate(ws.iter_rows(min_row=ws.min_row,
                                                     max_row=min(ws.min_row + 10, ws.max_row),
                                                     max_col=min(ws.max_column, 30)), start=1):
            row_data = []
            for cell in row:
                cell_info = {
                    'coord': cell.coordinate,
                    'value': str(cell.value)[:100] if cell.value else None,
                    'has_formula': str(cell.value).startswith('=') if cell.value else False
                }
                row_data.append(cell_info)

                # 수식이 있는 셀 추적
                if cell.value and str(cell.value).startswith('='):
                    sheet_info['formulas'].append({
                        'coord': cell.coordinate,
                        'formula': str(cell.value)[:200]
                    })
                    sheet_info['output_cells'].append(cell.coordinate)
                elif cell.value and not str(cell.value).startswith('='):
                    # 값이 있고 수식이 아닌 경우 잠재적 입력 셀
                    sheet_info['input_cells'].append(cell.coordinate)

            sheet_info['sample_data'].append(row_data)

        # 수식 샘플 제한
        sheet_info['formulas'] = sheet_info['formulas'][:50]
        sheet_info['input_cells'] = sheet_info['input_cells'][:100]
        sheet_info['output_cells'] = sheet_info['output_cells'][:100]

        result['sheets'][sheet_name] = sheet_info

    wb.close()
    return result

def analyze_xls_file(file_path):
    """xls 파일 분석 (xlrd 사용)"""
    import xlrd

    result = {}
    wb = xlrd.open_workbook(file_path)

    result['filename'] = os.path.basename(file_path)
    result['filepath'] = file_path
    result['sheet_count'] = wb.nsheets
    result['sheet_names'] = wb.sheet_names()
    result['sheets'] = {}

    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        sheet_info = {
            'dimensions': f'A1:{xlrd.colname(ws.ncols-1)}{ws.nrows}' if ws.nrows > 0 and ws.ncols > 0 else 'Empty',
            'min_row': 1,
            'max_row': ws.nrows,
            'min_col': 1,
            'max_col': ws.ncols,
            'sample_data': []
        }

        # 샘플 데이터
        for row_idx in range(min(10, ws.nrows)):
            row_data = []
            for col_idx in range(min(30, ws.ncols)):
                cell = ws.cell(row_idx, col_idx)
                row_data.append({
                    'coord': f'{xlrd.colname(col_idx)}{row_idx+1}',
                    'value': str(cell.value)[:100] if cell.value else None,
                    'type': cell.ctype
                })
            sheet_info['sample_data'].append(row_data)

        result['sheets'][sheet_name] = sheet_info

    return result

def main():
    base_path = r'C:\Users\rando\repos\reAA'
    output_file = os.path.join(base_path, 'excel_analysis_result.json')

    files_to_analyze = [
        os.path.join(base_path, '경주대종천하류보20250501_unlocked.xlsm'),
        os.path.join(base_path, '경주_대종천_하류보.xls'),
        os.path.join(base_path, '경주대종천하류보20250501.xls')
    ]

    all_results = []

    for file_path in files_to_analyze:
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            continue

        print(f"Analyzing: {os.path.basename(file_path)}")

        try:
            if file_path.endswith('.xlsm') or file_path.endswith('.xlsx'):
                result = analyze_xlsm_file(file_path)
            else:
                result = analyze_xls_file(file_path)
            all_results.append(result)
        except Exception as e:
            print(f"Error analyzing {file_path}: {e}")
            all_results.append({'filename': os.path.basename(file_path), 'error': str(e)})

    # UTF-8로 JSON 저장
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)

    print(f"\nAnalysis complete. Results saved to: {output_file}")

if __name__ == '__main__':
    main()
