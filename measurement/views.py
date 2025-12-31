from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods, require_GET, require_POST
from django.views.decorators.csrf import csrf_exempt
from io import BytesIO
from datetime import datetime, timedelta
import json
import math


def measurement_list(request):
    """측정 목록"""
    from .models import MeasurementSession

    # DB에서 측정 세션 조회 (모든 사용자)
    sessions = MeasurementSession.objects.order_by('-updated_at')[:50]

    measurements = []
    for s in sessions:
        measurements.append({
            'id': s.pk,
            'station_name': s.station_name or '미지정',
            'river_name': '',
            'measurement_date': s.measurement_date.strftime('%Y-%m-%d') if s.measurement_date else '-',
            'discharge': s.estimated_discharge or 0,
            'uncertainty': '-',
            'status': 'completed' if s.estimated_discharge else 'draft',
            'session_number': s.session_number,
            'updated_at': s.updated_at.strftime('%Y-%m-%d %H:%M') if s.updated_at else '',
        })

    return render(request, 'measurement/list.html', {
        'measurements': measurements,
    })


def measurement_new(request):
    """새 측정 생성 마법사"""
    if request.method == 'POST':
        # 폼 처리 (추후 구현)
        return redirect('measurement:data_input')
    return render(request, 'measurement/new.html')


def measurement_detail(request, pk):
    """측정 상세 (결과 보기) - DB에서 데이터 로드"""
    from .models import MeasurementSession

    session = get_object_or_404(MeasurementSession, pk=pk)

    # 세션에 rows_data가 있으면 계산 수행, 없으면 저장된 값 사용
    if session.rows_data:
        calibration = session.calibration_data or {'a': 0.0012, 'b': 0.2534}
        result_data = calculate_discharge(session.rows_data, calibration)
    else:
        # rows_data가 없으면 저장된 값으로 표시
        setup = session.setup_data or {}
        result_data = {
            'discharge': session.estimated_discharge or 0,
            'area': session.total_area or 0,
            'width': session.total_width or 0,
            'avg_depth': session.max_depth or 0,
            'avg_velocity': setup.get('final_avg_velocity', 0),
            'max_velocity': 0,
            'min_velocity': 0,
            'uncertainty': setup.get('final_uncertainty', 0),
            'uncertainty_abs': 0,
            'verticals': [],
        }
        if result_data['discharge'] and result_data['uncertainty']:
            result_data['uncertainty_abs'] = round(
                result_data['discharge'] * result_data['uncertainty'] / 100, 3
            )

    # 세션 정보 추가
    result_data['session_id'] = session.pk
    result_data['station_name'] = session.station_name or ''
    result_data['measurement_date'] = session.measurement_date.strftime('%Y-%m-%d') if session.measurement_date else ''
    result_data['setup_data'] = session.setup_data or {}

    # measurement 객체도 전달 (템플릿 호환성)
    result_data['measurement'] = {
        'id': session.pk,
        'station_name': session.station_name,
        'date': result_data['measurement_date'],
    }

    return render(request, 'measurement/result.html', result_data)


def measurement_edit(request, pk):
    """측정 수정 - 세션 ID를 전달하여 data_input으로 이동"""
    return redirect(f'/measurement/data-input/?session_id={pk}')


def measurement_delete(request, pk):
    """측정 삭제"""
    from .models import MeasurementSession

    if request.method == 'DELETE':
        try:
            session = MeasurementSession.objects.get(pk=pk)
            session.delete()
            return HttpResponse(status=200)
        except MeasurementSession.DoesNotExist:
            return HttpResponse(status=404)
    return HttpResponse(status=405)


def data_input(request):
    """데이터 입력 그리드"""
    from django.conf import settings
    return render(request, 'measurement/data_input.html', {
        'debug': settings.DEBUG,
    })


def result(request):
    """계산 결과"""
    from .models import MeasurementSession

    # GET 요청: session_id로 데이터 로드 후 계산 (권장 방식)
    session_id = request.GET.get('session_id')
    if session_id:
        try:
            session = MeasurementSession.objects.get(pk=session_id)
            rows = session.rows_data or []
            calibration = session.calibration_data or {'a': 0.0012, 'b': 0.2534}

            # 계산 수행
            result_data = calculate_discharge(rows, calibration)

            # 세션 정보 추가
            result_data['session_id'] = session.pk
            result_data['station_name'] = session.station_name or ''
            result_data['measurement_date'] = session.measurement_date.strftime('%Y-%m-%d') if session.measurement_date else ''
            result_data['setup_data'] = session.setup_data or {}

            # 계산 결과를 세션에 저장 (자동 저장)
            session.estimated_discharge = result_data['discharge']
            session.total_area = result_data['area']
            session.total_width = result_data['width']
            session.max_depth = result_data['avg_depth']
            if not session.setup_data:
                session.setup_data = {}
            session.setup_data['final_discharge'] = result_data['discharge']
            session.setup_data['final_uncertainty'] = result_data['uncertainty']
            session.setup_data['final_avg_velocity'] = result_data['avg_velocity']
            session.save()

            return render(request, 'measurement/result.html', result_data)
        except MeasurementSession.DoesNotExist:
            return render(request, 'measurement/result.html', {'error': '세션을 찾을 수 없습니다.'})

    # POST 요청: 레거시 방식 (deprecated - 추후 제거 예정)
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            rows = data.get('rows', [])
            calibration = data.get('calibration', {'a': 0.0012, 'b': 0.2534})
            meters = data.get('meters', None)  # 다중 유속계 목록

            # 세션 정보 추출
            session_id = data.get('session_id')
            station_name = data.get('station_name', '')
            measurement_date = data.get('measurement_date', '')
            setup_data = data.get('setup_data', {})

            # 계산 수행 (다중 유속계 지원)
            result_data = calculate_discharge(rows, calibration, meters)

            # 세션 정보 추가
            result_data['session_id'] = session_id
            result_data['station_name'] = station_name
            result_data['measurement_date'] = measurement_date
            result_data['setup_data'] = setup_data

            return render(request, 'measurement/result.html', result_data)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return render(request, 'measurement/result.html')


def calculate_discharge(rows, calibration, meters=None):
    """
    유량 계산 (ISO 748 중앙단면법)

    Args:
        rows: 측선 데이터 리스트
        calibration: 기본 검정계수 (단일 유속계 사용 시)
        meters: 유속계 목록 (다중 유속계 사용 시) [{'id': 'M1', 'a': 0.0012, 'b': 0.2534, 'uncertainty': 1.0}, ...]

    Returns dict with:
    - discharge: 총 유량 (m³/s)
    - area: 총 단면적 (m²)
    - avg_velocity: 평균유속 (m/s)
    - uncertainty: 불확실도 (%)
    - verticals: 각 측선별 상세 결과
    """
    # 기본 검정계수
    default_a = calibration.get('a', 0.0012)
    default_b = calibration.get('b', 0.2534)

    # 유속계 목록을 딕셔너리로 변환 (빠른 조회용)
    meters_dict = {}
    if meters:
        for m in meters:
            meters_dict[m.get('id')] = {
                'a': m.get('a', default_a),
                'b': m.get('b', default_b),
                'uncertainty': m.get('uncertainty', 1.0)
            }

    def get_calibration(meter_id):
        """유속계별 검정계수 조회"""
        if meter_id and meter_id in meters_dict:
            m = meters_dict[meter_id]
            return m['a'], m['b'], m['uncertainty']
        return default_a, default_b, 1.0

    def calc_velocity(n, t, a, b):
        """프로펠러 유속계 유속 계산: V = a + b * (N/T)"""
        try:
            n = float(n) if n else 0
            t = float(t) if t else 0
            if n == 0 or t == 0:
                return 0
            return a + b * (n / t)
        except (ValueError, TypeError):
            return 0

    # 각 측선별 유속 계산
    verticals = []
    meter_uncertainties_list = []  # 유속계별 검정 불확실도 수집

    for i, row in enumerate(rows):
        method = row.get('method', '1')
        meter_id = row.get('meter_id')
        velocity = 0

        # 측선별 유속계 검정계수 조회
        a, b, meter_uncertainty = get_calibration(meter_id)
        meter_uncertainties_list.append(meter_uncertainty)

        if method == '1':  # 1점법
            velocity = calc_velocity(row.get('n_06d'), row.get('t_06d'), a, b)
        elif method == '2':  # 2점법
            v02 = calc_velocity(row.get('n_02d'), row.get('t_02d'), a, b)
            v08 = calc_velocity(row.get('n_08d'), row.get('t_08d'), a, b)
            if v02 and v08:
                velocity = (v02 + v08) / 2
        elif method == '3':  # 3점법
            v02 = calc_velocity(row.get('n_02d'), row.get('t_02d'), a, b)
            v06 = calc_velocity(row.get('n_06d'), row.get('t_06d'), a, b)
            v08 = calc_velocity(row.get('n_08d'), row.get('t_08d'), a, b)
            if v02 and v06 and v08:
                velocity = (v02 + 2 * v06 + v08) / 4

        # 각도/index 보정 적용 (cos(θ) 또는 직접 index)
        # index_value: 0.6~1.0 범위 (1.0이면 보정 없음)
        index_value = float(row.get('index_value', 1.0) or 1.0)
        angle_deg = float(row.get('angle_deg', 0) or 0)

        # 각도가 입력되었으면 cos(θ) 계산, 아니면 index_value 사용
        if angle_deg > 0:
            angle_rad = math.radians(angle_deg)
            index_value = math.cos(angle_rad)

        if velocity > 0:
            velocity = velocity * index_value

        verticals.append({
            'id': i + 1,
            'distance': float(row.get('distance', 0) or 0),
            'depth': float(row.get('depth', 0) or 0),
            'meter_id': meter_id,
            'angle_deg': angle_deg,
            'index_value': round(index_value, 4),
            'velocity': velocity,
            'method': method,
            'area': 0,
            'discharge': 0,
            'ratio': 0
        })

    # 중앙단면법으로 유량 계산
    total_discharge = 0
    total_area = 0

    for i in range(1, len(verticals)):
        prev = verticals[i - 1]
        curr = verticals[i]

        width = curr['distance'] - prev['distance']
        avg_depth = (curr['depth'] + prev['depth']) / 2
        avg_velocity = (curr['velocity'] + prev['velocity']) / 2

        section_area = width * avg_depth
        section_discharge = section_area * avg_velocity

        # 현재 측선에 값 저장 (중앙단면법에서 각 구간의 값)
        curr['width'] = width
        curr['area'] = section_area
        curr['discharge'] = section_discharge

        total_area += section_area
        total_discharge += section_discharge

    # 비율 계산
    for v in verticals:
        if total_discharge > 0:
            v['ratio'] = (v['discharge'] / total_discharge) * 100

    # 평균유속
    avg_velocity = total_discharge / total_area if total_area > 0 else 0

    # 최대/최소 유속 (0이 아닌 값 중에서)
    velocities = [v['velocity'] for v in verticals if v['velocity'] > 0]
    max_velocity = max(velocities) if velocities else 0
    min_velocity = min(velocities) if velocities else 0

    # 하폭
    distances = [v['distance'] for v in verticals]
    total_width = max(distances) - min(distances) if distances else 0

    # 평균수심
    avg_depth = total_area / total_width if total_width > 0 else 0

    # ============================================
    # ISO 748 불확실도 계산 (완전 구현)
    # ============================================
    n_verticals = len([v for v in verticals if v['velocity'] > 0])
    methods = [v['method'] for v in verticals if v['method'] in ['1', '2', '3']]

    # ----- X1Q: 랜덤 불확실도 (Random Uncertainty) -----

    # Xm: 측선수 불확실도 (ISO 748 Table 1 기반)
    # 측선수가 많을수록 불확실도 감소
    if n_verticals >= 25:
        Xm = 0.5
    elif n_verticals >= 20:
        Xm = 1.0
    elif n_verticals >= 15:
        Xm = 1.5
    elif n_verticals >= 10:
        Xm = 2.0
    elif n_verticals >= 5:
        Xm = 3.0
    else:
        Xm = 5.0

    # Xp: 측점수 불확실도 (측정법별)
    # 1점법: 15%, 2점법: 7%, 3점법: 6.33%
    method_xp = {'1': 15.0, '2': 7.0, '3': 6.33}
    if methods:
        Xp_values = [method_xp.get(m, 15.0) for m in methods]
        Xp = sum(Xp_values) / len(Xp_values)
    else:
        Xp = 15.0

    # Xc: 유속계 검정 불확실도 (각 측선별 유속계의 불확실도 평균)
    valid_meter_uncertainties = [u for i, u in enumerate(meter_uncertainties_list)
                                  if i < len(verticals) and verticals[i]['velocity'] > 0]
    if valid_meter_uncertainties:
        Xc = sum(valid_meter_uncertainties) / len(valid_meter_uncertainties)
    else:
        Xc = 1.0

    # Xe: 측점 불확실도 (노출시간, 유속에 따라 변동)
    # 간략화: 평균유속에 따라 결정
    if avg_velocity >= 1.0:
        Xe = 1.5  # 고유속: 낮은 불확실도
    elif avg_velocity >= 0.5:
        Xe = 2.0  # 중유속
    elif avg_velocity >= 0.2:
        Xe = 3.0  # 저유속
    else:
        Xe = 5.0  # 극저유속: 높은 불확실도

    # X1Q: 랜덤 불확실도 합성
    # X1Q = sqrt(Xm² + (Xe² + Xp² + Xc²) / n)
    if n_verticals > 0:
        random_component = (Xe**2 + Xp**2 + Xc**2) / n_verticals
        X1Q = math.sqrt(Xm**2 + random_component)
    else:
        X1Q = math.sqrt(Xm**2 + Xe**2 + Xp**2 + Xc**2)

    # ----- X2Q: 계통 불확실도 (Systematic Uncertainty) -----
    X2b = 0.5   # 폭 측정 불확실도
    X2d = 0.5   # 수심 측정 불확실도
    X2c = 1.0   # 기타 계통 불확실도
    X2Q = math.sqrt(X2b**2 + X2d**2 + X2c**2)  # = 1.22%

    # ----- 총 불확실도 -----
    # XQ = sqrt(X1Q² + X2Q²)
    combined_uncertainty = math.sqrt(X1Q**2 + X2Q**2)

    # 확장불확실도 (95%, k=2)
    expanded_uncertainty = combined_uncertainty * 2
    uncertainty_abs = total_discharge * expanded_uncertainty / 100

    # 유량등급 결정 (ISO 748)
    if expanded_uncertainty <= 5:
        quality_grade = 'E'  # Excellent
    elif expanded_uncertainty <= 8:
        quality_grade = 'G'  # Good
    elif expanded_uncertainty <= 10:
        quality_grade = 'F'  # Fair
    else:
        quality_grade = 'P'  # Poor

    # 유량 일 단위 환산 (m³/s → m³/d)
    discharge_daily = round(total_discharge * 86400, 0)

    return {
        'discharge': round(total_discharge, 3),
        'discharge_daily': f"{discharge_daily:,.0f}",  # 천 단위 구분 포맷
        'area': round(total_area, 2),
        'avg_velocity': round(avg_velocity, 3),
        'max_velocity': round(max_velocity, 3),
        'min_velocity': round(min_velocity, 3),
        'width': round(total_width, 1),
        'avg_depth': round(avg_depth, 2),
        'uncertainty': round(expanded_uncertainty, 1),
        'uncertainty_abs': round(uncertainty_abs, 3),
        # 불확실도 상세 구성요소
        'Xe': round(Xe, 2),
        'Xp': round(Xp, 2),
        'Xc': round(Xc, 2),
        'Xm': round(Xm, 2),
        'X1Q': round(X1Q, 2),  # 랜덤 불확실도
        'X2Q': round(X2Q, 2),  # 계통 불확실도
        'combined_uncertainty': round(combined_uncertainty, 2),
        'quality_grade': quality_grade,
        'verticals': verticals,
        'n_verticals': n_verticals,
    }


def pre_uncertainty(request):
    """사전 불확실도 분석"""
    return render(request, 'measurement/pre_uncertainty.html')


def meters(request):
    """유속계 관리"""
    return render(request, 'measurement/meters.html')


def export_excel(request):
    """Excel 다운로드 - session_id로 DB에서 데이터 로드"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from .models import MeasurementSession

    session_id = request.GET.get('session_id')

    if session_id:
        try:
            session = MeasurementSession.objects.get(pk=session_id)
            calibration = session.calibration_data or {'a': 0.0012, 'b': 0.2534}

            if session.rows_data:
                result = calculate_discharge(session.rows_data, calibration)
                rows = []
                for v in result.get('verticals', []):
                    rows.append({
                        'no': v['id'],
                        'distance': v['distance'],
                        'depth': v['depth'],
                        'velocity': round(v['velocity'], 3),
                        'area': round(v.get('area', 0), 3),
                        'discharge': round(v.get('discharge', 0), 3),
                        'ratio': round(v.get('ratio', 0), 1),
                    })
                data = {
                    'station_name': session.station_name or '미지정',
                    'date': session.measurement_date.strftime('%Y-%m-%d') if session.measurement_date else '-',
                    'discharge': result['discharge'],
                    'uncertainty': result['uncertainty'],
                    'area': result['area'],
                    'avg_velocity': result['avg_velocity'],
                    'rows': rows,
                }
            else:
                setup = session.setup_data or {}
                data = {
                    'station_name': session.station_name or '미지정',
                    'date': session.measurement_date.strftime('%Y-%m-%d') if session.measurement_date else '-',
                    'discharge': session.estimated_discharge or 0,
                    'uncertainty': setup.get('final_uncertainty', 0),
                    'area': session.total_area or 0,
                    'avg_velocity': setup.get('final_avg_velocity', 0),
                    'rows': [],
                }
        except MeasurementSession.DoesNotExist:
            return HttpResponse('세션을 찾을 수 없습니다.', status=404)
    else:
        return HttpResponse('session_id 파라미터가 필요합니다.', status=400)

    wb = Workbook()
    ws = wb.active
    ws.title = "유량측정 결과"

    # 스타일 정의
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # 제목
    ws['A1'] = "유량측정 결과 보고서"
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')

    # 기본 정보
    ws['A3'] = "지점명"
    ws['B3'] = data['station_name']
    ws['A4'] = "측정일"
    ws['B4'] = data['date']
    ws['D3'] = "유량 (m³/s)"
    ws['E3'] = data['discharge']
    ws['D4'] = "불확실도 (%)"
    ws['E4'] = f"± {data['uncertainty']}"

    # 데이터 테이블 헤더 (7행부터)
    headers = ['측선', '거리 (m)', '수심 (m)', '유속 (m/s)', '단면적 (m²)', '유량 (m³/s)', '비율 (%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 데이터 행
    for row_idx, row_data in enumerate(data['rows'], 8):
        ws.cell(row=row_idx, column=1, value=row_data['no']).border = thin_border
        ws.cell(row=row_idx, column=2, value=row_data['distance']).border = thin_border
        ws.cell(row=row_idx, column=3, value=row_data['depth']).border = thin_border
        ws.cell(row=row_idx, column=4, value=row_data['velocity']).border = thin_border
        ws.cell(row=row_idx, column=5, value=row_data['area']).border = thin_border
        ws.cell(row=row_idx, column=6, value=row_data['discharge']).border = thin_border
        ws.cell(row=row_idx, column=7, value=row_data['ratio']).border = thin_border
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).alignment = center_align

    # 합계 행
    sum_row = 8 + len(data['rows'])
    ws.cell(row=sum_row, column=1, value='합계').font = Font(bold=True)
    ws.merge_cells(f'A{sum_row}:D{sum_row}')
    ws.cell(row=sum_row, column=5, value=data['area']).font = Font(bold=True)
    ws.cell(row=sum_row, column=6, value=data['discharge']).font = Font(bold=True)
    ws.cell(row=sum_row, column=7, value=100.0).font = Font(bold=True)
    for col in range(1, 8):
        ws.cell(row=sum_row, column=col).border = thin_border
        ws.cell(row=sum_row, column=col).alignment = center_align

    # 열 너비 조정
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12

    # 응답 생성
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"measurement_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_pdf(request):
    """PDF 출력 - session_id로 DB에서 데이터 로드"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from .models import MeasurementSession
    import os
    from django.conf import settings

    # 한글 폰트 등록
    font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'NotoSansKR-Regular.ttf')
    if os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont('NotoSansKR', font_path))
            KOREAN_FONT = 'NotoSansKR'
        except:
            KOREAN_FONT = 'Helvetica'
    else:
        KOREAN_FONT = 'Helvetica'

    session_id = request.GET.get('session_id')

    if session_id:
        try:
            session = MeasurementSession.objects.get(pk=session_id)
            calibration = session.calibration_data or {'a': 0.0012, 'b': 0.2534}

            if session.rows_data:
                result = calculate_discharge(session.rows_data, calibration)
                rows = []
                for v in result.get('verticals', []):
                    rows.append([
                        v['id'],
                        v['distance'],
                        round(v['depth'], 2),
                        round(v['velocity'], 3),
                        round(v.get('area', 0), 3),
                        round(v.get('discharge', 0), 3),
                        f"{round(v.get('ratio', 0), 1)}%",
                    ])
                data = {
                    'station_name': session.station_name or '미지정',
                    'date': session.measurement_date.strftime('%Y-%m-%d') if session.measurement_date else '-',
                    'discharge': result['discharge'],
                    'uncertainty': result['uncertainty'],
                    'area': result['area'],
                    'avg_velocity': result['avg_velocity'],
                    'rows': rows,
                }
            else:
                setup = session.setup_data or {}
                data = {
                    'station_name': session.station_name or '미지정',
                    'date': session.measurement_date.strftime('%Y-%m-%d') if session.measurement_date else '-',
                    'discharge': session.estimated_discharge or 0,
                    'uncertainty': setup.get('final_uncertainty', 0),
                    'area': session.total_area or 0,
                    'avg_velocity': setup.get('final_avg_velocity', 0),
                    'rows': [],
                }
        except MeasurementSession.DoesNotExist:
            return HttpResponse('세션을 찾을 수 없습니다.', status=404)
    else:
        return HttpResponse('session_id 파라미터가 필요합니다.', status=400)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontName=KOREAN_FONT,
        fontSize=16,
        spaceAfter=12,
        alignment=1  # center
    )
    normal_style = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontName=KOREAN_FONT,
        fontSize=10,
        spaceAfter=6
    )
    # 한글 폰트가 적용된 Heading2 스타일
    heading2_style = ParagraphStyle(
        'Heading2Korean',
        parent=styles['Heading2'],
        fontName=KOREAN_FONT,
        fontSize=12,
        spaceAfter=6
    )

    # 제목
    elements.append(Paragraph("Discharge Measurement Report", title_style))
    elements.append(Paragraph("유량측정 결과 보고서", title_style))
    elements.append(Spacer(1, 10*mm))

    # 기본 정보 테이블
    info_data = [
        ['Station', data['station_name'], 'Date', data['date']],
        ['Discharge', f"{data['discharge']} m³/s", 'Uncertainty', f"± {data['uncertainty']}%"],
        ['Area', f"{data['area']} m²", 'Avg. Velocity', f"{data['avg_velocity']} m/s"],
    ]
    info_table = Table(info_data, colWidths=[35*mm, 55*mm, 35*mm, 45*mm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.Color(0.9, 0.9, 0.9)),
        ('BACKGROUND', (2, 0), (2, -1), colors.Color(0.9, 0.9, 0.9)),
        ('FONTNAME', (0, 0), (-1, -1), KOREAN_FONT),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 10*mm))

    # 상세 데이터 테이블
    elements.append(Paragraph("Detailed Results by Vertical", heading2_style))
    elements.append(Spacer(1, 5*mm))

    table_data = [['No.', 'Distance(m)', 'Depth(m)', 'Velocity(m/s)', 'Area(m²)', 'Discharge(m³/s)', 'Ratio']]
    table_data.extend(data['rows'])
    table_data.append(['Total', '', '', '', data['area'], data['discharge'], '100%'])

    detail_table = Table(table_data, colWidths=[15*mm, 25*mm, 22*mm, 28*mm, 22*mm, 32*mm, 20*mm])
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.27, 0.45, 0.77)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, -1), KOREAN_FONT),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(detail_table)
    elements.append(Spacer(1, 10*mm))

    # 불확실도 정보
    elements.append(Paragraph("Uncertainty Analysis (ISO 748)", heading2_style))
    elements.append(Spacer(1, 5*mm))

    unc_data = [
        ['Component', 'Value', 'Description'],
        ['Xe', '1.8%', 'Limited number of verticals'],
        ['Xp', '2.1%', 'Depth measurement uncertainty'],
        ['Xc', '1.0%', 'Calibration uncertainty'],
        ['Xm', '2.8%', 'Velocity measurement uncertainty'],
        ['u(Q)', '4.2%', 'Combined standard uncertainty (RSS)'],
    ]
    unc_table = Table(unc_data, colWidths=[30*mm, 25*mm, 80*mm])
    unc_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.27, 0.45, 0.77)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, -1), KOREAN_FONT),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (1, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.85, 0.92, 0.98)),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(unc_table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"measurement_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def rating_curve_list(request):
    """수위-유량곡선 목록"""
    from .models import RatingCurve

    curves_qs = RatingCurve.objects.select_related('station').all()

    # 템플릿용 데이터 변환
    curves = []
    for curve in curves_qs:
        curves.append({
            'id': curve.pk,
            'station_name': curve.station.name,
            'year': curve.year,
            'curve_type': curve.get_curve_type_display(),
            'h_range': f'{curve.h_min:.2f} ≤ h ≤ {curve.h_max:.2f}',
            'equation': curve.get_equation_display(),
            'r_squared': curve.r_squared,
        })

    return render(request, 'measurement/rating_curve_list.html', {
        'curves': curves,
    })


def rating_curve_detail(request, pk):
    """수위-유량곡선 상세"""
    from .models import RatingCurve, HQDataPoint
    from django.shortcuts import get_object_or_404

    curve = get_object_or_404(RatingCurve.objects.select_related('station'), pk=pk)

    # 실측 데이터 조회
    data_points = HQDataPoint.objects.filter(rating_curve=curve).order_by('measured_date')
    data_points_list = [
        {
            'date': dp.measured_date.strftime('%Y-%m-%d'),
            'stage': dp.stage,
            'discharge': dp.discharge,
        }
        for dp in data_points
    ]

    curve_data = {
        'id': curve.pk,
        'station_name': curve.station.name,
        'year': curve.year,
        'curve_type': curve.curve_type,
        'curve_type_display': curve.get_curve_type_display(),
        'h_min': curve.h_min,
        'h_max': curve.h_max,
        'coef_a': curve.coef_a,
        'coef_b': curve.coef_b,
        'coef_h0': curve.coef_h0,
        'r_squared': curve.r_squared,
        'equation': curve.get_equation_display(),
        'data_points': data_points_list,
        'note': curve.note,
    }

    return render(request, 'measurement/rating_curve_detail.html', {
        'curve': curve_data,
    })


def rating_curve_new(request):
    """새 수위-유량곡선 생성"""
    return render(request, 'measurement/rating_curve_new.html')


@require_http_methods(["POST"])
def fit_rating_curve(request):
    """수위-유량 데이터로 곡선 피팅 (AJAX)"""
    import numpy as np
    from scipy.optimize import curve_fit

    try:
        data = json.loads(request.body)
        h_values = np.array(data.get('h', []))
        q_values = np.array(data.get('q', []))
        h0_fixed = data.get('h0', None)  # 고정 영점수위 (선택)

        if len(h_values) < 3 or len(q_values) < 3:
            return JsonResponse({'error': '최소 3개 이상의 데이터가 필요합니다.'}, status=400)

        # 멱함수: Q = a * (h - h0)^b
        if h0_fixed is not None:
            # h0 고정
            def power_func(h, a, b):
                return a * np.power(h - h0_fixed, b)

            popt, pcov = curve_fit(power_func, h_values, q_values, p0=[1, 1.5], maxfev=10000)
            a, b = popt
            h0 = h0_fixed
        else:
            # h0도 피팅
            def power_func_h0(h, a, b, h0):
                return a * np.power(np.maximum(h - h0, 1e-10), b)

            # 초기값 추정
            h0_init = min(h_values) * 0.9
            popt, pcov = curve_fit(power_func_h0, h_values, q_values, p0=[1, 1.5, h0_init], maxfev=10000)
            a, b, h0 = popt

        # 예측값 및 통계
        if h0_fixed is not None:
            q_pred = power_func(h_values, a, b)
        else:
            q_pred = power_func_h0(h_values, a, b, h0)

        # R² 계산
        ss_res = np.sum((q_values - q_pred) ** 2)
        ss_tot = np.sum((q_values - np.mean(q_values)) ** 2)
        r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else 0

        # RMSE 계산
        rmse = np.sqrt(np.mean((q_values - q_pred) ** 2))

        # 곡선 데이터 생성 (그래프용)
        h_curve = np.linspace(min(h_values), max(h_values), 50)
        if h0_fixed is not None:
            q_curve = power_func(h_curve, a, b)
        else:
            q_curve = power_func_h0(h_curve, a, b, h0)

        return JsonResponse({
            'success': True,
            'coefficients': {
                'a': round(float(a), 6),
                'b': round(float(b), 6),
                'h0': round(float(h0), 6),
            },
            'statistics': {
                'r_squared': round(float(r_squared), 4),
                'rmse': round(float(rmse), 6),
            },
            'equation': f"Q = {a:.4f}(h - {h0:.4f})^{b:.4f}",
            'h_range': {
                'min': round(float(min(h_values)), 3),
                'max': round(float(max(h_values)), 3),
            },
            'curve_points': {
                'h': h_curve.tolist(),
                'q': q_curve.tolist(),
            },
            'fitted_values': q_pred.tolist(),
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_POST
def api_rating_curve_save(request):
    """수위-유량곡선 저장 API"""
    from .models import Station, RatingCurve, HQDataPoint

    try:
        data = json.loads(request.body)

        station_name = data.get('station_name', '').strip()
        year = int(data.get('year', datetime.now().year))
        curve_type = data.get('curve_type', 'open')
        coef_a = float(data.get('coef_a', 0))
        coef_b = float(data.get('coef_b', 0))
        coef_h0 = float(data.get('coef_h0', 0))
        h_min = float(data.get('h_min', 0))
        h_max = float(data.get('h_max', 0))
        r_squared = data.get('r_squared')
        rmse = data.get('rmse')
        data_points = data.get('data_points', [])

        if not station_name:
            return JsonResponse({'error': '지점명을 입력하세요.'}, status=400)

        # 사용자 정보
        user = request.user if request.user.is_authenticated else None

        # Station 찾거나 생성
        station, created = Station.objects.get_or_create(
            name=station_name,
            defaults={'user': user}
        )

        # RatingCurve 생성
        rating_curve = RatingCurve.objects.create(
            user=user,
            station=station,
            year=year,
            curve_type=curve_type,
            h_min=h_min,
            h_max=h_max,
            coef_a=coef_a,
            coef_b=coef_b,
            coef_h0=coef_h0,
            r_squared=r_squared,
            rmse=rmse,
        )

        # 실측 데이터 저장
        for dp in data_points:
            if dp.get('h') and dp.get('q'):
                HQDataPoint.objects.create(
                    station=station,
                    rating_curve=rating_curve,
                    measured_date=dp.get('date') or datetime.now().date(),
                    stage=float(dp['h']),
                    discharge=float(dp['q']),
                )

        return JsonResponse({
            'success': True,
            'curve_id': rating_curve.pk,
            'message': '저장되었습니다.',
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def timeseries_list(request):
    """시계열 데이터 관리"""
    from .models import Station, WaterLevelTimeSeries, RatingCurve
    from django.db.models import Count, Min, Max

    # DB에서 수위 시계열 데이터가 있는 관측소 조회
    stations_with_data = WaterLevelTimeSeries.objects.values('station').annotate(
        data_count=Count('id'),
        first_time=Min('timestamp'),
        last_time=Max('timestamp'),
    )

    stations = []
    for item in stations_with_data:
        station = Station.objects.get(pk=item['station'])
        has_rating = RatingCurve.objects.filter(station=station).exists()

        first_time = item['first_time']
        last_time = item['last_time']
        period = f"{first_time.strftime('%Y-%m-%d')} ~ {last_time.strftime('%Y-%m-%d')}" if first_time and last_time else '-'

        stations.append({
            'id': station.pk,
            'name': station.name,
            'has_rating_curve': has_rating,
            'data_count': item['data_count'],
            'last_update': last_time.strftime('%Y-%m-%d %H:%M') if last_time else '-',
            'period': period,
        })

    return render(request, 'measurement/timeseries_list.html', {
        'stations': stations,
    })


def timeseries_upload(request):
    """수위 데이터 업로드"""
    from .models import Station, RatingCurve

    if request.method == 'POST':
        # CSV 파싱 및 저장 로직 (추후 구현)
        return JsonResponse({'success': True, 'message': '업로드 완료'})

    # 관측소 및 H-Q 곡선 목록
    stations = Station.objects.all().order_by('name')
    rating_curves = RatingCurve.objects.select_related('station').order_by('station__name', '-year')

    return render(request, 'measurement/timeseries_upload.html', {
        'stations': stations,
        'rating_curves': rating_curves,
    })


@require_GET
def api_stations_search(request):
    """관측소 검색 API"""
    from .models import Station

    query = request.GET.get('q', '').strip()
    limit = int(request.GET.get('limit', 20))

    stations = Station.objects.all()

    if query:
        stations = stations.filter(name__icontains=query)

    stations = stations.order_by('name')[:limit]

    results = []
    for s in stations:
        results.append({
            'id': s.pk,
            'name': s.name,
            'river_name': s.river_name or '',
        })

    return JsonResponse({
        'stations': results,
        'query': query,
    })


@require_GET
def api_rating_curves_by_station(request, station_id):
    """관측소별 H-Q 곡선 목록 API"""
    from .models import RatingCurve

    curves = RatingCurve.objects.filter(station_id=station_id).order_by('-year', 'curve_type')

    results = []
    for c in curves:
        results.append({
            'id': c.pk,
            'year': c.year,
            'curve_type': c.curve_type,
            'curve_type_display': c.get_curve_type_display(),
            'equation': c.get_equation_display(),
            'h_range': f'{c.h_min} ≤ h ≤ {c.h_max}',
        })

    return JsonResponse({
        'rating_curves': results,
        'station_id': station_id,
    })


def timeseries_detail(request, station_id):
    """관측소별 시계열 상세 - DB에서 데이터 로드"""
    from .models import Station, WaterLevelTimeSeries, RatingCurve

    station = get_object_or_404(Station, pk=station_id)

    # 수위 시계열 데이터 조회 (최신순 1000개)
    water_levels = WaterLevelTimeSeries.objects.filter(
        station=station
    ).order_by('-timestamp')[:1000]

    total_count = WaterLevelTimeSeries.objects.filter(station=station).count()

    # Rating Curve 조회 (가장 최신)
    rating_curve = RatingCurve.objects.filter(station=station).order_by('-year').first()

    # 시계열 데이터 변환
    data = []
    for wl in water_levels:
        discharge = None
        if rating_curve and wl.stage is not None:
            # Q = a * (h - h0) ^ b
            h_diff = wl.stage - rating_curve.coef_h0
            if h_diff > 0:
                discharge = round(rating_curve.coef_a * (h_diff ** rating_curve.coef_b), 4)

        data.append({
            'timestamp': wl.timestamp.strftime('%Y-%m-%d %H:%M'),
            'stage': wl.stage,
            'discharge': discharge,
        })

    # Rating Curve 수식 생성
    rating_curve_eq = None
    if rating_curve:
        rating_curve_eq = f"Q = {rating_curve.coef_a}(h - {rating_curve.coef_h0})^{rating_curve.coef_b}"

    station_data = {
        'id': station.pk,
        'name': station.name,
        'rating_curve': rating_curve_eq,
    }

    return render(request, 'measurement/timeseries_detail.html', {
        'station': station_data,
        'data': data[:100],  # 최근 100개만 템플릿에 전달
        'total_count': total_count,
    })


@require_http_methods(["POST"])
def generate_discharge_series(request):
    """Rating Curve 적용하여 유량 시계열 생성 (AJAX)"""
    from .models import Station, WaterLevelTimeSeries, RatingCurve, DischargeTimeSeries

    try:
        data = json.loads(request.body)
        station_id = data.get('station_id')
        rating_curve_id = data.get('rating_curve_id')

        if not station_id or not rating_curve_id:
            return JsonResponse({'error': 'station_id와 rating_curve_id가 필요합니다.'}, status=400)

        # 관측소 및 Rating Curve 조회
        station = get_object_or_404(Station, pk=station_id)
        rating_curve = get_object_or_404(RatingCurve, pk=rating_curve_id)

        # 수위 시계열 조회
        water_levels = WaterLevelTimeSeries.objects.filter(station=station)

        if not water_levels.exists():
            return JsonResponse({'error': '수위 시계열 데이터가 없습니다.'}, status=400)

        # 기존 유량 시계열 삭제 (같은 station + rating_curve 조합)
        DischargeTimeSeries.objects.filter(
            station=station,
            rating_curve=rating_curve
        ).delete()

        # 유량 계산 및 저장
        created_count = 0
        extrapolated_count = 0

        discharge_records = []
        for wl in water_levels:
            if wl.stage is None:
                continue

            # Q = a * (h - h0) ^ b
            h_diff = wl.stage - rating_curve.coef_h0

            if h_diff <= 0:
                # 수위가 h0 이하면 유량 0
                discharge = 0
                quality_flag = 'suspect'
            else:
                discharge = rating_curve.coef_a * (h_diff ** rating_curve.coef_b)

                # 외삽 여부 확인
                if wl.stage < rating_curve.h_min or wl.stage > rating_curve.h_max:
                    quality_flag = 'extrapolated'
                    extrapolated_count += 1
                else:
                    quality_flag = 'good'

            discharge_records.append(DischargeTimeSeries(
                station=station,
                timestamp=wl.timestamp,
                stage=wl.stage,
                discharge=round(discharge, 4),
                rating_curve=rating_curve,
                quality_flag=quality_flag,
            ))
            created_count += 1

        # 일괄 생성
        DischargeTimeSeries.objects.bulk_create(discharge_records, batch_size=1000)

        return JsonResponse({
            'success': True,
            'message': f'유량 시계열 생성 완료 ({created_count}개)',
            'count': created_count,
            'extrapolated': extrapolated_count,
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def baseflow_list(request):
    """기저유출 분석 목록"""
    from .models import BaseflowAnalysis

    # DB에서 분석 목록 조회
    db_analyses = BaseflowAnalysis.objects.select_related('station').order_by('-created_at')

    analyses = []
    for a in db_analyses:
        analyses.append({
            'id': a.pk,
            'station_name': a.station.name if a.station else 'Unknown',
            'period': f'{a.start_date} ~ {a.end_date}',
            'method': a.get_method_display(),
            'bfi': a.bfi,
            'total_runoff': a.total_runoff,
            'baseflow': a.baseflow,
            'created_at': a.created_at.strftime('%Y-%m-%d') if a.created_at else '',
        })

    return render(request, 'measurement/baseflow_list.html', {
        'analyses': analyses,
    })


def baseflow_new(request):
    """새 기저유출 분석"""
    from hydro.services import STATION_DATABASE
    from .models import Station

    # URL 파라미터에서 내부 관측소 ID 확인
    internal_station_id = request.GET.get('station')
    internal_station = None

    if internal_station_id:
        try:
            internal_station = Station.objects.get(pk=internal_station_id)
        except Station.DoesNotExist:
            pass

    # 내부 관측소 목록 (시계열 데이터가 있는 것만)
    from .models import WaterLevelTimeSeries
    internal_stations = Station.objects.filter(
        pk__in=WaterLevelTimeSeries.objects.values('station_id').distinct()
    )

    return render(request, 'measurement/baseflow_new.html', {
        'hrfco_stations': STATION_DATABASE,
        'internal_stations': internal_stations,
        'selected_internal_station': internal_station,
    })


@require_GET
def api_hrfco_discharge(request):
    """HRFCO API에서 유량 데이터 가져오기"""
    from hydro.services import get_waterlevel_history
    from datetime import datetime, timedelta

    station_code = request.GET.get('station_code')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not station_code:
        return JsonResponse({'error': '관측소 코드가 필요합니다.'}, status=400)

    try:
        # 기본 기간: 최근 1년
        if not end_date:
            end_dt = datetime.now()
        else:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')

        if not start_date:
            start_dt = end_dt - timedelta(days=365)
        else:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')

        # HRFCO API 호출
        data = get_waterlevel_history(station_code, start_dt, end_dt)

        # 일별 평균 유량으로 집계 (기저유출 분석용)
        daily_data = {}
        for item in data:
            if item.get('fw') is not None:  # 유량 데이터가 있는 경우만
                date_key = item['datetime'].strftime('%Y-%m-%d') if item.get('datetime') else None
                if date_key:
                    if date_key not in daily_data:
                        daily_data[date_key] = []
                    daily_data[date_key].append(item['fw'])

        # 일평균 계산
        discharge_series = []
        dates = []
        for date_key in sorted(daily_data.keys()):
            values = daily_data[date_key]
            avg_discharge = sum(values) / len(values)
            dates.append(date_key)
            discharge_series.append(round(avg_discharge, 4))

        return JsonResponse({
            'success': True,
            'station_code': station_code,
            'dates': dates,
            'discharge': discharge_series,
            'count': len(discharge_series),
            'start_date': dates[0] if dates else None,
            'end_date': dates[-1] if dates else None,
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_GET
def api_internal_discharge(request):
    """내부 관측소(WaterLevelTimeSeries)에서 유량 데이터 가져오기"""
    from .models import Station, WaterLevelTimeSeries, RatingCurve
    from django.db.models import Avg
    from django.db.models.functions import TruncDate

    station_id = request.GET.get('station_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not station_id:
        return JsonResponse({'error': '관측소 ID가 필요합니다.'}, status=400)

    try:
        station = Station.objects.get(pk=station_id)

        # Rating Curve 조회 (가장 최신)
        rating_curve = RatingCurve.objects.filter(station=station).order_by('-year').first()
        if not rating_curve:
            return JsonResponse({'error': '해당 관측소에 Rating Curve가 없습니다.'}, status=400)

        # 기간 설정
        from datetime import datetime, timedelta
        if not end_date:
            end_dt = datetime.now()
        else:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')

        if not start_date:
            start_dt = end_dt - timedelta(days=365)
        else:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')

        # 수위 시계열 조회 (일평균)
        from django.utils import timezone
        start_aware = timezone.make_aware(start_dt) if timezone.is_naive(start_dt) else start_dt
        end_aware = timezone.make_aware(end_dt) if timezone.is_naive(end_dt) else end_dt

        daily_stages = WaterLevelTimeSeries.objects.filter(
            station=station,
            timestamp__gte=start_aware,
            timestamp__lte=end_aware,
            quality_flag='good'
        ).annotate(
            date=TruncDate('timestamp')
        ).values('date').annotate(
            avg_stage=Avg('stage')
        ).order_by('date')

        # 수위 → 유량 변환 (Rating Curve 적용)
        dates = []
        discharge_series = []
        a, b, h0 = rating_curve.coef_a, rating_curve.coef_b, rating_curve.coef_h0

        for item in daily_stages:
            h_diff = item['avg_stage'] - h0
            if h_diff > 0:
                discharge = a * (h_diff ** b)
            else:
                discharge = 0

            dates.append(item['date'].strftime('%Y-%m-%d'))
            discharge_series.append(round(discharge, 4))

        return JsonResponse({
            'success': True,
            'station_id': station_id,
            'station_name': station.name,
            'rating_curve': f"Q = {a}(h - {h0})^{b}",
            'dates': dates,
            'discharge': discharge_series,
            'count': len(discharge_series),
            'start_date': dates[0] if dates else None,
            'end_date': dates[-1] if dates else None,
        })

    except Station.DoesNotExist:
        return JsonResponse({'error': '관측소를 찾을 수 없습니다.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def baseflow_detail(request, pk):
    """기저유출 분석 상세"""
    from .models import BaseflowAnalysis

    try:
        # DB에서 분석 결과 조회
        analysis_obj = BaseflowAnalysis.objects.select_related('station').get(pk=pk)
        daily_results = analysis_obj.daily_results.all().order_by('date')

        # 일별 데이터 변환
        daily_data = []
        for d in daily_results:
            direct = d.total_discharge - d.baseflow if d.total_discharge and d.baseflow else 0
            daily_data.append({
                'date': d.date.strftime('%Y-%m-%d'),
                'total': round(d.total_discharge, 3) if d.total_discharge else 0,
                'baseflow': round(d.baseflow, 3) if d.baseflow else 0,
                'direct': round(max(0, direct), 3),
            })

        # 분석 결과 객체 (템플릿에서 .pk 접근 가능)
        class AnalysisWrapper:
            def __init__(self, obj):
                self.pk = obj.pk
                self.id = obj.pk
                self.station_name = obj.station.name if obj.station else 'Unknown'
                self.start_date = obj.start_date
                self.end_date = obj.end_date
                self.method = obj.method
                self.method_display = obj.get_method_display()
                self.alpha = obj.alpha
                self.bfi_max = obj.bfi_max
                self.total_runoff = obj.total_runoff
                self.baseflow = obj.baseflow
                self.direct_runoff = obj.direct_runoff
                self.bfi = obj.bfi

        analysis = AnalysisWrapper(analysis_obj)

        return render(request, 'measurement/baseflow_detail.html', {
            'analysis': analysis,
            'daily_data': daily_data,
        })

    except BaseflowAnalysis.DoesNotExist:
        from django.http import Http404
        raise Http404("분석 결과를 찾을 수 없습니다.")


@require_http_methods(["POST"])
def run_baseflow_analysis(request):
    """기저유출 분석 실행 (AJAX)"""
    import numpy as np

    try:
        data = json.loads(request.body)
        method = data.get('method', 'lyne_hollick')
        alpha = float(data.get('alpha', 0.925))
        discharge = np.array(data.get('discharge', []))

        if len(discharge) < 30:
            return JsonResponse({'error': '최소 30일 이상의 데이터가 필요합니다.'}, status=400)

        # Lyne-Hollick 디지털 필터
        if method == 'lyne_hollick':
            baseflow = np.zeros_like(discharge)
            baseflow[0] = discharge[0] * 0.5

            for i in range(1, len(discharge)):
                # b(i) = α * b(i-1) + (1-α)/2 * (Q(i) + Q(i-1))
                baseflow[i] = alpha * baseflow[i-1] + (1 - alpha) / 2 * (discharge[i] + discharge[i-1])
                baseflow[i] = min(baseflow[i], discharge[i])

        # Eckhardt 필터
        elif method == 'eckhardt':
            bfi_max = float(data.get('bfi_max', 0.80))
            baseflow = np.zeros_like(discharge)
            baseflow[0] = discharge[0] * bfi_max

            for i in range(1, len(discharge)):
                baseflow[i] = ((1 - bfi_max) * alpha * baseflow[i-1] +
                               (1 - alpha) * bfi_max * discharge[i]) / (1 - alpha * bfi_max)
                baseflow[i] = min(baseflow[i], discharge[i])

        else:
            return JsonResponse({'error': '지원하지 않는 분석 방법입니다.'}, status=400)

        # 결과 계산
        direct_runoff = discharge - baseflow
        total_sum = float(np.sum(discharge))
        baseflow_sum = float(np.sum(baseflow))
        bfi = baseflow_sum / total_sum if total_sum > 0 else 0

        return JsonResponse({
            'success': True,
            'baseflow': baseflow.tolist(),
            'direct_runoff': direct_runoff.tolist(),
            'statistics': {
                'total_runoff': round(total_sum, 2),
                'baseflow': round(baseflow_sum, 2),
                'direct_runoff': round(float(np.sum(direct_runoff)), 2),
                'bfi': round(bfi, 4),
            }
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["POST"])
def save_baseflow_analysis(request):
    """기저유출 분석 결과 저장"""
    from .models import Station, BaseflowAnalysis, BaseflowDaily
    from datetime import datetime

    # 로그인 체크
    if not request.user.is_authenticated:
        return JsonResponse({'error': '로그인이 필요합니다.'}, status=401)

    try:
        data = json.loads(request.body)

        # 필수 데이터
        station_id = data.get('station_id')
        station_name = data.get('station_name', '')
        method = data.get('method', 'lyne_hollick')
        alpha = float(data.get('alpha', 0.925))
        bfi_max = float(data.get('bfi_max', 0.80)) if data.get('bfi_max') else None
        start_date = data.get('start_date')
        end_date = data.get('end_date')

        # 분석 결과
        statistics = data.get('statistics', {})
        daily_data = data.get('daily_data', [])  # [{date, total, baseflow, direct}, ...]

        # 관측소 조회 또는 생성
        station = None
        if station_id:
            try:
                station = Station.objects.get(id=station_id)
            except Station.DoesNotExist:
                pass

        if not station and station_name:
            station, _ = Station.objects.get_or_create(
                name=station_name,
                defaults={'description': '자동 생성'}
            )

        if not station:
            return JsonResponse({'error': '관측소 정보가 필요합니다.'}, status=400)

        # BaseflowAnalysis 생성
        analysis = BaseflowAnalysis.objects.create(
            user=request.user,  # 사용자 연결
            station=station,
            start_date=datetime.strptime(start_date, '%Y-%m-%d').date(),
            end_date=datetime.strptime(end_date, '%Y-%m-%d').date(),
            method=method,
            alpha=alpha,
            bfi_max=bfi_max,
            total_runoff=statistics.get('total_runoff'),
            baseflow=statistics.get('baseflow'),
            direct_runoff=statistics.get('direct_runoff'),
            bfi=statistics.get('bfi'),
        )

        # 활동 로그 기록
        from core.tracking import log_activity
        log_activity(
            user=request.user,
            action_type='save',
            detail=f'기저유출 분석 저장: {station.name} ({start_date}~{end_date})',
            related_object=analysis,
            request=request,
            extra_data={'method': method, 'bfi': statistics.get('bfi')}
        )

        # 일별 데이터 저장
        daily_objects = []
        for d in daily_data:
            daily_objects.append(BaseflowDaily(
                analysis=analysis,
                date=datetime.strptime(d['date'], '%Y-%m-%d').date(),
                total_discharge=d['total'],
                baseflow=d['baseflow'],
                direct_runoff=d['direct'],
            ))

        if daily_objects:
            BaseflowDaily.objects.bulk_create(daily_objects)

        return JsonResponse({
            'success': True,
            'analysis_id': analysis.id,
            'message': f'분석 결과가 저장되었습니다. (ID: {analysis.id})'
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_GET
def export_baseflow_pdf(request, pk):
    """기저유출 분석 PDF 리포트 다운로드"""
    from django.http import HttpResponse
    from .models import BaseflowAnalysis
    from .pdf_service import generate_baseflow_report
    import urllib.parse

    try:
        analysis = BaseflowAnalysis.objects.select_related('station').get(pk=pk)
        daily_data = list(analysis.daily_results.all().order_by('date'))

        # PDF 생성
        pdf_buffer = generate_baseflow_report(analysis, daily_data)

        # 안전한 파일명 생성
        station_name = analysis.station.name if analysis.station else 'unknown'
        safe_filename = urllib.parse.quote(f"baseflow_{station_name}_{analysis.start_date}.pdf")

        # 응답 생성
        response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{safe_filename}"

        # 활동 로그 기록
        if request.user.is_authenticated:
            from core.tracking import log_activity
            log_activity(
                user=request.user,
                action_type='export_pdf',
                detail=f'기저유출 PDF 다운로드: {station_name}',
                related_object=analysis,
                request=request
            )

        return response

    except BaseflowAnalysis.DoesNotExist:
        return HttpResponse('Analysis not found', status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f'Error generating PDF: {str(e)}', status=500)


# ============================================
# 측정 데이터 자동저장 및 히스토리
# ============================================

@require_http_methods(["POST"])
def api_measurement_autosave(request):
    """측정 데이터 자동저장 API"""
    from .models import MeasurementSession

    try:
        data = json.loads(request.body)

        session_id = data.get('session_id')  # 기존 세션 업데이트 시
        station_name = data.get('station_name', '')
        measurement_date = data.get('measurement_date')
        session_number = int(data.get('session_number', 1))
        rows_data = data.get('rows', [])
        calibration_data = data.get('calibration', {})
        setup_data = data.get('setup_data', {})  # 하천명, 기상, 측정자 등
        estimated_discharge = data.get('estimated_discharge')
        total_width = data.get('total_width')
        max_depth = data.get('max_depth')
        total_area = data.get('total_area')

        # 날짜 파싱
        parsed_date = None
        if measurement_date:
            try:
                parsed_date = datetime.strptime(measurement_date, '%Y-%m-%d').date()
            except:
                pass

        # 사용자 또는 세션키로 식별
        user = request.user if request.user.is_authenticated else None
        session_key = request.session.session_key or ''
        if not session_key and not user:
            request.session.create()
            session_key = request.session.session_key

        # 기존 세션 업데이트 또는 새로 생성
        session = None
        if session_id:
            try:
                if user:
                    session = MeasurementSession.objects.get(pk=session_id, user=user)
                else:
                    session = MeasurementSession.objects.get(pk=session_id, session_key=session_key)
            except MeasurementSession.DoesNotExist:
                session = None

        if session:
            # 기존 세션 업데이트
            session.station_name = station_name
            session.measurement_date = parsed_date
            session.session_number = session_number
            session.rows_data = rows_data
            session.calibration_data = calibration_data
            session.setup_data = setup_data
            session.estimated_discharge = estimated_discharge
            session.total_width = total_width
            session.max_depth = max_depth
            session.total_area = total_area
            session.save()
        else:
            # 새 세션 생성
            session = MeasurementSession.objects.create(
                user=user,
                session_key=session_key,
                station_name=station_name,
                measurement_date=parsed_date,
                session_number=session_number,
                rows_data=rows_data,
                calibration_data=calibration_data,
                setup_data=setup_data,
                estimated_discharge=estimated_discharge,
                total_width=total_width,
                max_depth=max_depth,
                total_area=total_area,
            )

        return JsonResponse({
            'success': True,
            'session_id': session.pk,
            'updated_at': session.updated_at.isoformat(),
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_POST
def api_result_save(request):
    """결과 페이지에서 최종 저장 API

    중복 저장 방지:
    - session_id가 있으면 해당 세션만 업데이트
    - session_id가 없으면 동일 데이터 중복 체크 후 저장
    - 트랜잭션 + select_for_update로 race condition 방지
    """
    from .models import MeasurementSession
    from django.db import transaction

    try:
        data = json.loads(request.body)

        session_id = data.get('session_id')
        station_name = data.get('station_name', '')
        measurement_date = data.get('measurement_date')
        discharge = data.get('discharge')
        area = data.get('area')
        avg_velocity = data.get('avg_velocity')
        uncertainty = data.get('uncertainty')

        # 날짜 파싱
        parsed_date = None
        if measurement_date:
            try:
                parsed_date = datetime.strptime(measurement_date, '%Y-%m-%d').date()
            except:
                pass

        # session_id가 있으면 해당 세션만 업데이트 (중복 생성 방지)
        if session_id:
            with transaction.atomic():
                try:
                    session = MeasurementSession.objects.select_for_update().get(pk=session_id)
                    # 기존 세션 업데이트 - 결과값 저장
                    session.station_name = station_name or session.station_name
                    if parsed_date:
                        session.measurement_date = parsed_date
                    session.estimated_discharge = discharge
                    session.total_area = area
                    if not session.setup_data:
                        session.setup_data = {}
                    session.setup_data['final_discharge'] = discharge
                    session.setup_data['final_uncertainty'] = uncertainty
                    session.setup_data['final_avg_velocity'] = avg_velocity
                    session.save()

                    return JsonResponse({
                        'success': True,
                        'session_id': session.pk,
                        'message': '결과가 저장되었습니다.',
                    })
                except MeasurementSession.DoesNotExist:
                    pass  # session_id가 유효하지 않으면 아래에서 중복 체크 후 처리

        # 사용자 또는 세션키로 식별
        user = request.user if request.user.is_authenticated else None
        session_key = request.session.session_key or ''
        if not session_key and not user:
            request.session.create()
            session_key = request.session.session_key

        # 트랜잭션 내에서 중복 체크 및 생성 (race condition 방지)
        with transaction.atomic():
            # 중복 체크: 동일 관측소 + 날짜 + 유량값이 있으면 기존 세션 반환
            if station_name and parsed_date and discharge:
                existing = MeasurementSession.objects.select_for_update().filter(
                    station_name=station_name,
                    measurement_date=parsed_date,
                    estimated_discharge=discharge
                ).first()

                if existing:
                    return JsonResponse({
                        'success': True,
                        'session_id': existing.pk,
                        'message': '이미 저장된 데이터입니다.',
                        'duplicate': True,
                    })

            # 새 세션 생성
            session = MeasurementSession.objects.create(
                user=user,
                session_key=session_key,
                station_name=station_name,
                measurement_date=parsed_date,
                estimated_discharge=discharge,
                total_area=area,
                setup_data={
                    'final_discharge': discharge,
                    'final_uncertainty': uncertainty,
                    'final_avg_velocity': avg_velocity,
                }
            )

        return JsonResponse({
            'success': True,
            'session_id': session.pk,
            'message': '결과가 저장되었습니다.',
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_GET
def api_measurement_history(request):
    """측정 데이터 히스토리 목록 API"""
    from .models import MeasurementSession

    limit = int(request.GET.get('limit', 20))

    # 사용자 또는 세션키로 필터
    user = request.user if request.user.is_authenticated else None
    session_key = request.session.session_key or ''

    if user:
        sessions = MeasurementSession.objects.filter(user=user)
    elif session_key:
        sessions = MeasurementSession.objects.filter(session_key=session_key)
    else:
        # 비로그인이고 세션도 없으면 전체 최근 데이터 (개발용)
        sessions = MeasurementSession.objects.all()

    sessions = sessions.order_by('-updated_at')[:limit]

    history = []
    for s in sessions:
        history.append({
            'id': s.pk,
            'station_name': s.station_name or '미지정',
            'measurement_date': s.measurement_date.strftime('%Y-%m-%d') if s.measurement_date else None,
            'session_number': s.session_number,
            'estimated_discharge': s.estimated_discharge,
            'row_count': len(s.rows_data) if s.rows_data else 0,
            'updated_at': s.updated_at.strftime('%Y-%m-%d %H:%M'),
            'created_at': s.created_at.strftime('%Y-%m-%d %H:%M'),
        })

    return JsonResponse({
        'history': history,
        'count': len(history),
    })


@require_GET
def api_measurement_load(request, session_id):
    """저장된 측정 데이터 불러오기 API"""
    from .models import MeasurementSession

    try:
        # 개발 모드에서는 모든 세션 접근 허용
        session = MeasurementSession.objects.get(pk=session_id)

        return JsonResponse({
            'success': True,
            'session_id': session.pk,
            'station_name': session.station_name,
            'measurement_date': session.measurement_date.strftime('%Y-%m-%d') if session.measurement_date else None,
            'session_number': session.session_number,
            'rows': session.rows_data,
            'calibration': session.calibration_data,
            'setup_data': session.setup_data,
            'estimated_discharge': session.estimated_discharge,
            'total_width': session.total_width,
            'max_depth': session.max_depth,
            'total_area': session.total_area,
            'updated_at': session.updated_at.isoformat(),
        })

    except MeasurementSession.DoesNotExist:
        return JsonResponse({'error': '세션을 찾을 수 없습니다.'}, status=404)


@require_http_methods(["DELETE"])
def api_measurement_delete(request, session_id):
    """측정 데이터 세션 삭제 API"""
    from .models import MeasurementSession

    try:
        # 개발 모드에서는 모든 세션 삭제 허용
        session = MeasurementSession.objects.get(pk=session_id)
        session.delete()

        return JsonResponse({
            'success': True,
            'message': '삭제되었습니다.',
        })

    except MeasurementSession.DoesNotExist:
        return JsonResponse({'error': '세션을 찾을 수 없습니다.'}, status=404)


@csrf_exempt
def api_create_mock_data(request):
    """개발용: 모의 데이터 생성 API"""
    from django.core.management import call_command
    from io import StringIO

    try:
        out = StringIO()
        clean = request.GET.get('clean', '').lower() == 'true'
        call_command('create_mock_data', clean=clean, stdout=out)
        output = out.getvalue()

        return JsonResponse({
            'success': True,
            'message': '모의 데이터 생성 완료',
            'output': output,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def api_create_sample_upstream_downstream(request):
    """샘플 데이터 생성: 같은 지점/날짜에 상류+하류 데이터"""
    from datetime import date
    from .models import MeasurementSession

    try:
        # 상류 데이터
        upstream = MeasurementSession.objects.create(
            station_name='[샘플] 경주 대종천',
            measurement_date=date(2025, 12, 22),
            session_number=1,
            setup_data={
                'location_desc': '보상류 50m',
                'river_name': '대종천',
                'weather': '맑음',
                'measurer': '테스트'
            },
            rows_data=[
                {'distance': 0, 'depth': 0.2, 'velocity': 0.1, 'method': '1'},
                {'distance': 2, 'depth': 0.8, 'velocity': 0.45, 'method': '1'},
                {'distance': 4, 'depth': 1.2, 'velocity': 0.65, 'method': '1'},
                {'distance': 6, 'depth': 1.5, 'velocity': 0.72, 'method': '1'},
                {'distance': 8, 'depth': 1.3, 'velocity': 0.58, 'method': '1'},
                {'distance': 10, 'depth': 0.7, 'velocity': 0.35, 'method': '1'},
                {'distance': 12, 'depth': 0.2, 'velocity': 0.1, 'method': '1'}
            ],
            estimated_discharge=4.25,
            total_width=12.0,
            total_area=7.2,
            mean_velocity=0.59,
            uncertainty=8.5,
            quality_grade='G'
        )
        upstream.calculate_analysis_results()
        upstream.save()

        # 하류 데이터
        downstream = MeasurementSession.objects.create(
            station_name='[샘플] 경주 대종천',
            measurement_date=date(2025, 12, 22),
            session_number=2,
            setup_data={
                'location_desc': '보하류 100m',
                'river_name': '대종천',
                'weather': '맑음',
                'measurer': '테스트'
            },
            rows_data=[
                {'distance': 0, 'depth': 0.3, 'velocity': 0.15, 'method': '1'},
                {'distance': 3, 'depth': 1.0, 'velocity': 0.55, 'method': '1'},
                {'distance': 6, 'depth': 1.8, 'velocity': 0.85, 'method': '1'},
                {'distance': 9, 'depth': 2.0, 'velocity': 0.92, 'method': '1'},
                {'distance': 12, 'depth': 1.6, 'velocity': 0.78, 'method': '1'},
                {'distance': 15, 'depth': 0.9, 'velocity': 0.48, 'method': '1'},
                {'distance': 18, 'depth': 0.3, 'velocity': 0.12, 'method': '1'}
            ],
            estimated_discharge=8.65,
            total_width=18.0,
            total_area=11.5,
            mean_velocity=0.75,
            uncertainty=7.2,
            quality_grade='G'
        )
        downstream.calculate_analysis_results()
        downstream.save()

        return JsonResponse({
            'success': True,
            'message': '상류/하류 샘플 데이터 생성 완료',
            'upstream_id': upstream.id,
            'downstream_id': downstream.id
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ============================================
# 유속계 관리 API
# ============================================

def _format_date(date_value):
    """날짜 값을 문자열로 변환 (이미 문자열이거나 None인 경우 처리)"""
    if date_value is None:
        return None
    if isinstance(date_value, str):
        return date_value
    try:
        return date_value.strftime('%Y-%m-%d')
    except (AttributeError, TypeError):
        return str(date_value) if date_value else None


def _meter_to_dict_list(meter):
    """Meter 객체를 목록용 딕셔너리로 변환"""
    return {
        'id': meter.id,
        'meter_id': meter.meter_id,
        'type': meter.meter_type,
        'a': meter.coef_a,
        'b': meter.coef_b,
        'range': meter.range_display,
        'range_min': meter.range_min,
        'range_max': meter.range_max,
        'uncertainty': meter.uncertainty,
        'calibrationDate': _format_date(meter.calibration_date),
        'calibrationExpiry': _format_date(getattr(meter, 'calibration_expiry', None)),
        'calibrationOrg': meter.calibration_org,
        'status': meter.status,
    }


@require_GET
def api_meters_list(request):
    """유속계 목록 조회 API"""
    from .models import Meter

    meters = Meter.objects.all()
    data = [_meter_to_dict_list(m) for m in meters]

    return JsonResponse({'success': True, 'meters': data})


@csrf_exempt
@require_http_methods(["POST"])
def api_meters_create(request):
    """유속계 생성 API"""
    from .models import Meter
    from datetime import datetime
    import json

    try:
        data = json.loads(request.body)

        # 날짜 파싱
        calibration_date = None
        calibration_expiry = None
        if data.get('calibration_date'):
            try:
                calibration_date = datetime.strptime(data['calibration_date'], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                pass
        if data.get('calibration_expiry'):
            try:
                calibration_expiry = datetime.strptime(data['calibration_expiry'], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                pass

        meter = Meter.objects.create(
            meter_id=data.get('meter_id', ''),
            meter_type=data.get('type', 'propeller'),
            coef_a=float(data.get('a', 0)),
            coef_b=float(data.get('b', 1)),
            range_min=float(data.get('range_min', 0)),
            range_max=float(data.get('range_max', 6.0)),
            uncertainty=float(data.get('uncertainty', 1.0)),
            calibration_date=calibration_date,
            calibration_expiry=calibration_expiry,
            calibration_org=data.get('calibration_org', ''),
            status=data.get('status', 'valid'),
        )

        return JsonResponse({
            'success': True,
            'message': '유속계가 등록되었습니다.',
            'meter': _meter_to_dict_list(meter)
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["PUT"])
def api_meters_update(request, meter_id):
    """유속계 수정 API"""
    from .models import Meter
    from datetime import datetime
    import json

    try:
        meter = Meter.objects.get(pk=meter_id)
        data = json.loads(request.body)

        meter.meter_id = data.get('meter_id', meter.meter_id)
        meter.meter_type = data.get('type', meter.meter_type)
        meter.coef_a = float(data.get('a', meter.coef_a))
        meter.coef_b = float(data.get('b', meter.coef_b))
        meter.range_min = float(data.get('range_min', meter.range_min))
        meter.range_max = float(data.get('range_max', meter.range_max))
        meter.uncertainty = float(data.get('uncertainty', meter.uncertainty))
        if 'calibration_date' in data:
            if data['calibration_date']:
                try:
                    meter.calibration_date = datetime.strptime(data['calibration_date'], '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    meter.calibration_date = None
            else:
                meter.calibration_date = None
        if 'calibration_expiry' in data:
            if data['calibration_expiry']:
                try:
                    meter.calibration_expiry = datetime.strptime(data['calibration_expiry'], '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    meter.calibration_expiry = None
            else:
                meter.calibration_expiry = None
        if 'calibration_org' in data:
            meter.calibration_org = data['calibration_org']
        if 'status' in data:
            meter.status = data['status']

        meter.save()

        return JsonResponse({
            'success': True,
            'message': '유속계가 수정되었습니다.',
            'meter': _meter_to_dict_list(meter)
        })
    except Meter.DoesNotExist:
        return JsonResponse({'error': '유속계를 찾을 수 없습니다.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["DELETE"])
def api_meters_delete(request, meter_id):
    """유속계 삭제 API"""
    from .models import Meter

    try:
        meter = Meter.objects.get(pk=meter_id)
        meter.delete()

        return JsonResponse({
            'success': True,
            'message': '유속계가 삭제되었습니다.',
        })
    except Meter.DoesNotExist:
        return JsonResponse({'error': '유속계를 찾을 수 없습니다.'}, status=404)


# ============================================================
# 분석결과표 API
# ============================================================

def api_analysis_summary(request):
    """분석결과표 API - 측정 세션 종합 요약"""
    from .models import MeasurementSession

    # 필터 파라미터
    station_name = request.GET.get('station', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # 쿼리 생성 (로그인 여부와 관계없이 모든 데이터 표시)
    sessions = MeasurementSession.objects.all()

    if station_name:
        sessions = sessions.filter(station_name__icontains=station_name)

    if start_date:
        sessions = sessions.filter(measurement_date__gte=start_date)

    if end_date:
        sessions = sessions.filter(measurement_date__lte=end_date)

    # 측정일 기준 정렬
    sessions = sessions.order_by('measurement_date', 'station_name')

    # 분석 결과 계산 및 저장
    results = []
    for session in sessions:
        # 아직 계산되지 않은 경우 계산
        if session.wetted_perimeter is None and session.rows_data:
            session.calculate_analysis_results()
            session.save()

        results.append(session.to_summary_dict())

    return JsonResponse({
        'success': True,
        'count': len(results),
        'results': results,
    })


def api_analysis_recalculate(request, session_id):
    """특정 세션의 분석 결과 재계산"""
    from .models import MeasurementSession

    try:
        session = MeasurementSession.objects.get(pk=session_id)

        # 재계산
        session.calculate_analysis_results()
        session.save()

        return JsonResponse({
            'success': True,
            'result': session.to_summary_dict(),
        })
    except MeasurementSession.DoesNotExist:
        return JsonResponse({'error': '세션을 찾을 수 없습니다.'}, status=404)


def api_analysis_export(request):
    """분석결과표 Excel 다운로드 - 피벗 형식 (지역별 시트, 날짜별 열) + 차트"""
    import csv
    from collections import defaultdict
    from django.http import HttpResponse
    from .models import MeasurementSession

    # 차트 생성 함수
    def generate_chart_image(session):
        """세션 데이터로 단면/유속 차트 이미지 생성"""
        import traceback
        try:
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
            from io import BytesIO

            # 한글 폰트 설정 시도
            try:
                import matplotlib.font_manager as fm
                import os
                from django.conf import settings

                # 프로젝트 static 폴더의 NotoSansKR 폰트 우선 사용
                noto_font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'NotoSansKR-Regular.ttf')

                font_paths = [
                    noto_font_path,  # 프로젝트 내 한글 폰트 우선
                    '/usr/share/fonts/truetype/nanum/NanumGothic.ttf',
                    '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
                    'C:/Windows/Fonts/malgun.ttf',
                ]
                font_set = False
                for fp in font_paths:
                    try:
                        if os.path.exists(fp):
                            fm.fontManager.addfont(fp)
                            font_prop = fm.FontProperties(fname=fp)
                            plt.rcParams['font.family'] = font_prop.get_name()
                            font_set = True
                            print(f"Chart font set to: {fp}")
                            break
                    except Exception as e:
                        print(f"Font {fp} failed: {e}")
                        pass
                if not font_set:
                    plt.rcParams['font.family'] = 'DejaVu Sans'
                    print("Using fallback font: DejaVu Sans")
            except Exception as font_err:
                print(f"Font setup warning: {font_err}")
                plt.rcParams['font.family'] = 'DejaVu Sans'

            plt.rcParams['axes.unicode_minus'] = False

            rows = session.rows_data or []
            if not rows or len(rows) < 2:
                print(f"No rows data for session {session.id}")
                return None

            distances = [float(r.get('distance') or 0) for r in rows]
            depths = [-float(r.get('depth') or 0) for r in rows]
            velocities = [float(r.get('velocity') or 0) for r in rows]

            if not distances or max(depths) == min(depths) == 0:
                print(f"Invalid data for session {session.id}")
                return None

            fig, ax1 = plt.subplots(figsize=(5, 3), dpi=100)

            # 수심 (단면)
            ax1.fill_between(distances, depths, 0, alpha=0.3, color='#1e3a5f', label='Depth')
            ax1.plot(distances, depths, color='#1e3a5f', linewidth=2)
            ax1.set_xlabel('Distance (m)', fontsize=9)
            ax1.set_ylabel('Depth (m)', color='#1e3a5f', fontsize=9)
            ax1.tick_params(axis='y', labelcolor='#1e3a5f')
            min_depth = min(depths) if depths else -1
            ax1.set_ylim(min_depth * 1.2 if min_depth < 0 else -1, 0.1)
            ax1.axhline(y=0, color='#3b82f6', linewidth=1.5)
            ax1.grid(True, linestyle='--', alpha=0.5, color='gray')
            ax1.set_axisbelow(True)

            # 유속
            ax2 = ax1.twinx()
            ax2.plot(distances, velocities, color='#ef4444', linewidth=2, marker='o', markersize=4, label='Velocity')
            ax2.set_ylabel('Velocity (m/s)', color='#ef4444', fontsize=9)
            ax2.tick_params(axis='y', labelcolor='#ef4444')
            max_vel = max(velocities) if velocities else 1
            ax2.set_ylim(0, max_vel * 1.3 if max_vel > 0 else 1)
            ax2.grid(True, linestyle=':', alpha=0.3, color='#ef4444', axis='y')

            # 제목
            loc_desc = session.setup_data.get('location_desc', '') if session.setup_data else ''
            title = f"{session.station_name or 'Station'}"
            if loc_desc:
                title += f" ({loc_desc})"
            ax1.set_title(title, fontsize=10, fontweight='bold')

            # 범례
            lines1, labels1 = ax1.get_legend_handles_labels()
            lines2, labels2 = ax2.get_legend_handles_labels()
            ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper right', fontsize=8)

            plt.tight_layout()

            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', bbox_inches='tight', facecolor='white')
            plt.close(fig)
            img_buffer.seek(0)

            # 버퍼 크기 확인
            img_buffer.seek(0, 2)  # 끝으로 이동
            size = img_buffer.tell()
            img_buffer.seek(0)  # 다시 처음으로
            print(f"Chart generated for session {session.id}: {size} bytes")

            return img_buffer
        except Exception as e:
            print(f"Chart generation error for session {getattr(session, 'id', 'unknown')}: {e}")
            traceback.print_exc()
            return None

    # 필터
    station_name = request.GET.get('station', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    export_format = request.GET.get('format', 'csv')

    sessions = MeasurementSession.objects.all()

    if request.user.is_authenticated:
        sessions = sessions.filter(user=request.user)

    if station_name:
        sessions = sessions.filter(station_name__icontains=station_name)
    if start_date:
        sessions = sessions.filter(measurement_date__gte=start_date)
    if end_date:
        sessions = sessions.filter(measurement_date__lte=end_date)

    sessions = sessions.order_by('station_name', 'measurement_date')

    # Excel 피벗 형식
    if export_format == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image as XLImage

            wb = Workbook()
            wb.remove(wb.active)

            # 지역별 그룹화
            stations_data = defaultdict(list)
            for session in sessions:
                if session.wetted_perimeter is None and session.rows_data:
                    session.calculate_analysis_results()
                    session.save()
                stations_data[session.station_name or '미지정'].append(session)

            # 스타일
            header_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')

            # 항목 행 정의
            row_items = [
                ('수면폭', 'total_width', 'm', 3),
                ('단면적', 'total_area', 'm²', 2),
                ('윤변', 'wetted_perimeter', 'm', 3),
                ('동수반경', 'hydraulic_radius', 'm', 3),
                ('수위', 'stage', 'm', 3),
                ('평균유속', 'mean_velocity', 'm/s', 3),
                ('평균유량', 'estimated_discharge', 'm³/s', 3),
                ('유속측선수', 'velocity_verticals', '개', 0),
                ('불확실도', 'uncertainty', '%', 2),
                ('수위고도', 'stage_elevation', 'El.m', 2),
                ('유량조사 등급', 'quality_grade', '', 0),
            ]

            for stn_name, station_sessions in stations_data.items():
                sheet_name = stn_name[:31].replace('[', '').replace(']', '').replace('/', '-')
                ws = wb.create_sheet(title=sheet_name)

                # 날짜별, 위치별 그룹화
                date_location_data = defaultdict(dict)
                dates_set = set()
                locations_set = set()

                for s in station_sessions:
                    if s.measurement_date:
                        date_str = f"{s.measurement_date.month}/{s.measurement_date.day}/{s.measurement_date.year}"
                    else:
                        date_str = '미지정'

                    loc = s.setup_data.get('location_desc', '') if s.setup_data else ''
                    if '상류' in loc:
                        loc_key = '보 상류'
                    elif '하류' in loc:
                        loc_key = '보 하류'
                    else:
                        loc_key = loc or f'측정{s.session_number}'

                    dates_set.add(date_str)
                    locations_set.add(loc_key)
                    date_location_data[date_str][loc_key] = s

                dates = sorted(dates_set)
                locations = sorted(locations_set) or ['측정1']

                # 헤더
                ws.cell(row=1, column=1, value='항목').border = thin_border
                ws.cell(row=1, column=1).fill = header_fill
                ws.cell(row=2, column=1, value='구분').border = thin_border
                ws.cell(row=2, column=1).fill = header_fill

                col = 2
                for date in dates:
                    num_locs = len(locations)
                    ws.cell(row=1, column=col, value=date).border = thin_border
                    ws.cell(row=1, column=col).fill = header_fill
                    ws.cell(row=1, column=col).alignment = center_align
                    if num_locs > 1:
                        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + num_locs - 1)

                    for i, loc in enumerate(locations):
                        ws.cell(row=2, column=col + i, value=loc).border = thin_border
                        ws.cell(row=2, column=col + i).fill = header_fill
                        ws.cell(row=2, column=col + i).alignment = center_align
                    col += num_locs

                unit_col = col
                ws.cell(row=1, column=unit_col, value='단위').border = thin_border
                ws.cell(row=1, column=unit_col).fill = header_fill
                ws.merge_cells(start_row=1, start_column=unit_col, end_row=2, end_column=unit_col)

                # 데이터 행
                for row_idx, (item_name, attr, unit, decimals) in enumerate(row_items, start=3):
                    ws.cell(row=row_idx, column=1, value=item_name).border = thin_border
                    if item_name == '평균유량':
                        ws.cell(row=row_idx, column=1).fill = yellow_fill

                    col = 2
                    for date in dates:
                        for loc in locations:
                            session = date_location_data.get(date, {}).get(loc)
                            value = ''
                            if session and attr:
                                raw_value = getattr(session, attr, None)
                                if raw_value is not None:
                                    if decimals > 0 and isinstance(raw_value, (int, float)):
                                        value = round(raw_value, decimals)
                                    else:
                                        value = raw_value

                            cell = ws.cell(row=row_idx, column=col, value=value)
                            cell.border = thin_border
                            cell.alignment = center_align
                            if item_name == '평균유량':
                                cell.fill = yellow_fill
                            col += 1

                    ws.cell(row=row_idx, column=unit_col, value=unit).border = thin_border
                    ws.cell(row=row_idx, column=unit_col).alignment = center_align

                # 컬럼 너비
                ws.column_dimensions['A'].width = 14
                for i in range(2, unit_col + 1):
                    ws.column_dimensions[get_column_letter(i)].width = 12

                # 차트 이미지 삽입
                chart_start_row = len(row_items) + 5
                chart_count = 0

                for date in dates:
                    for loc in locations:
                        session = date_location_data.get(date, {}).get(loc)
                        if session:
                            img_buffer = generate_chart_image(session)
                            if img_buffer:
                                img = XLImage(img_buffer)
                                img.width = 400
                                img.height = 240
                                row_offset = (chart_count // 2) * 15
                                col_offset = (chart_count % 2) * 8
                                cell_ref = f"{get_column_letter(1 + col_offset)}{chart_start_row + row_offset}"
                                ws.add_image(img, cell_ref)
                                chart_count += 1

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="analysis_summary.xlsx"'
            wb.save(response)
            return response

        except ImportError as e:
            print(f"Import error: {e}")
            export_format = 'csv'

    # CSV 폴백
    headers = ['측정일', '지점명', '회차', '위치', '수위(m)', '수면폭(m)', '단면적(m²)',
               '윤변(m)', '동수반경(m)', '평균유속(m/s)', '유량(m³/s)', '유속측선수', '불확실도(%)', '등급']
    rows = []
    for session in sessions:
        rows.append([
            session.measurement_date.strftime('%Y-%m-%d') if session.measurement_date else '',
            session.station_name, session.session_number,
            session.setup_data.get('location_desc', '') if session.setup_data else '',
            round(session.stage, 2) if session.stage else '',
            round(session.total_width, 2) if session.total_width else '',
            round(session.total_area, 4) if session.total_area else '',
            round(session.wetted_perimeter, 4) if session.wetted_perimeter else '',
            round(session.hydraulic_radius, 4) if session.hydraulic_radius else '',
            round(session.mean_velocity, 3) if session.mean_velocity else '',
            round(session.estimated_discharge, 6) if session.estimated_discharge else '',
            session.velocity_verticals or '',
            round(session.uncertainty, 2) if session.uncertainty else '',
            session.quality_grade or '',
        ])

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="analysis_summary.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


def api_debug_chart_test(request):
    """차트 생성 디버깅용 API"""
    import traceback
    from .models import MeasurementSession

    result = {
        'matplotlib_version': None,
        'backend': None,
        'fonts_available': [],
        'sessions_count': 0,
        'chart_tests': []
    }

    try:
        import matplotlib
        result['matplotlib_version'] = matplotlib.__version__
        result['backend'] = matplotlib.get_backend()

        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import matplotlib.font_manager as fm

        # 사용 가능한 폰트 (일부만)
        fonts = [f.name for f in fm.fontManager.ttflist[:20]]
        result['fonts_available'] = fonts

        # 세션 테스트
        sessions = MeasurementSession.objects.all()[:3]
        result['sessions_count'] = MeasurementSession.objects.count()

        for session in sessions:
            test = {
                'id': session.id,
                'name': session.station_name,
                'rows_count': len(session.rows_data) if session.rows_data else 0,
                'chart_generated': False,
                'chart_size': 0,
                'error': None
            }

            try:
                rows = session.rows_data or []
                if rows and len(rows) >= 2:
                    from io import BytesIO

                    distances = [float(r.get('distance') or 0) for r in rows]
                    depths = [-float(r.get('depth') or 0) for r in rows]
                    velocities = [float(r.get('velocity') or 0) for r in rows]

                    fig, ax = plt.subplots(figsize=(4, 2), dpi=72)
                    ax.fill_between(distances, depths, 0, alpha=0.3)
                    ax.plot(distances, depths, linewidth=1)
                    ax.set_title(f"Test: {session.station_name[:20]}")
                    plt.tight_layout()

                    buf = BytesIO()
                    plt.savefig(buf, format='png')
                    plt.close(fig)

                    test['chart_generated'] = True
                    test['chart_size'] = buf.tell()
                else:
                    test['error'] = f"Insufficient rows: {len(rows)}"
            except Exception as e:
                test['error'] = str(e)
                test['traceback'] = traceback.format_exc()

            result['chart_tests'].append(test)

    except Exception as e:
        result['error'] = str(e)
        result['traceback'] = traceback.format_exc()

    return JsonResponse(result)


@require_POST
def api_parquet_import(request):
    """Parquet 파일을 JSON으로 변환하여 반환"""
    import pandas as pd

    if 'file' not in request.FILES:
        return JsonResponse({'error': '파일이 없습니다.'}, status=400)

    uploaded_file = request.FILES['file']

    if not uploaded_file.name.endswith('.parquet'):
        return JsonResponse({'error': 'Parquet 파일만 지원합니다.'}, status=400)

    try:
        # Parquet 읽기
        df = pd.read_parquet(BytesIO(uploaded_file.read()))

        # 컬럼명 정규화 (다양한 형식 지원)
        column_mapping = {
            '거리': 'distance', 'distance': 'distance', 'dist': 'distance',
            '수심': 'depth', 'depth': 'depth',
            'n_02d': 'n_02d', 'n02d': 'n_02d', 'N_02D': 'n_02d',
            't_02d': 't_02d', 't02d': 't_02d', 'T_02D': 't_02d',
            'n_06d': 'n_06d', 'n06d': 'n_06d', 'N_06D': 'n_06d',
            't_06d': 't_06d', 't06d': 't_06d', 'T_06D': 't_06d',
            'n_08d': 'n_08d', 'n08d': 'n_08d', 'N_08D': 'n_08d',
            't_08d': 't_08d', 't08d': 't_08d', 'T_08D': 't_08d',
        }

        # 컬럼명 변환
        df.columns = [column_mapping.get(c.lower().strip(), c.lower().strip()) for c in df.columns]

        # 필수 컬럼 확인
        if 'distance' not in df.columns or 'depth' not in df.columns:
            return JsonResponse({
                'error': f'필수 컬럼(distance, depth)이 없습니다. 현재 컬럼: {list(df.columns)}'
            }, status=400)

        # 데이터를 rows 형식으로 변환
        rows = []
        for idx, row in df.iterrows():
            depth = float(row.get('depth', 0) or 0)

            # 측정법 자동 결정
            if depth == 0:
                method = 'LEW' if idx == 0 else 'REW'
            elif depth < 0.6:
                method = '1'
            elif depth < 1.0:
                method = '2'
            else:
                method = '3'

            rows.append({
                'id': idx + 1,
                'distance': float(row.get('distance', 0) or 0),
                'depth': depth,
                'method': method,
                'n_02d': float(row['n_02d']) if 'n_02d' in row and pd.notna(row['n_02d']) else None,
                't_02d': float(row['t_02d']) if 't_02d' in row and pd.notna(row['t_02d']) else None,
                'n_06d': float(row['n_06d']) if 'n_06d' in row and pd.notna(row['n_06d']) else None,
                't_06d': float(row['t_06d']) if 't_06d' in row and pd.notna(row['t_06d']) else None,
                'n_08d': float(row['n_08d']) if 'n_08d' in row and pd.notna(row['n_08d']) else None,
                't_08d': float(row['t_08d']) if 't_08d' in row and pd.notna(row['t_08d']) else None,
                'velocity': 0
            })

        return JsonResponse({
            'success': True,
            'rows': rows,
            'columns': list(df.columns),
            'count': len(rows)
        })

    except Exception as e:
        return JsonResponse({'error': f'파일 처리 오류: {str(e)}'}, status=500)


def analysis_summary_view(request):
    """분석결과표 페이지"""
    return render(request, 'measurement/analysis_summary.html')


def batch_import(request):
    """CSV 일괄 가져오기 페이지"""
    return render(request, 'measurement/batch_import.html')


@csrf_exempt
@require_POST
def api_batch_import_csv(request):
    """CSV 일괄 가져오기 API

    여러 CSV 파일을 받아서 각각 MeasurementSession으로 저장
    파일명 형식: {하천명}_{날짜}_{위치}.csv (예: 대종천_20250501_보상류.csv)
    """
    from .models import MeasurementSession, Meter
    import re

    try:
        files = request.FILES.getlist('files')
        if not files:
            return JsonResponse({'error': 'CSV 파일을 선택해주세요.'}, status=400)

        # 유속계 정보 (기본값 또는 요청에서 받음)
        calibration_a = float(request.POST.get('calibration_a', 0.0116))
        calibration_b = float(request.POST.get('calibration_b', 0.2505))
        calibration = {'a': calibration_a, 'b': calibration_b}

        # 사용자/세션 식별
        user = request.user if request.user.is_authenticated else None
        session_key = request.session.session_key or ''
        if not session_key and not user:
            request.session.create()
            session_key = request.session.session_key

        results = []
        errors = []

        for file in files:
            filename = file.name
            try:
                # 파일명에서 정보 파싱: 하천명_날짜_위치.csv
                name_without_ext = filename.rsplit('.', 1)[0]
                parts = name_without_ext.split('_')

                river_name = ''
                measurement_date = None
                location = ''

                if len(parts) >= 3:
                    river_name = parts[0]  # 대종천
                    date_str = parts[1]    # 20250501
                    location = parts[2]    # 보상류

                    # 날짜 파싱
                    try:
                        measurement_date = datetime.strptime(date_str, '%Y%m%d').date()
                    except:
                        pass
                elif len(parts) == 2:
                    # 날짜_위치 형식
                    date_str = parts[0]
                    location = parts[1]
                    try:
                        measurement_date = datetime.strptime(date_str, '%Y%m%d').date()
                    except:
                        pass

                # station_name 생성
                station_name = f"{river_name} {location}".strip() if river_name else location

                # CSV 파싱
                content = file.read().decode('utf-8-sig')
                lines = content.strip().split('\n')

                # 헤더 확인
                start_index = 0
                first_line = lines[0].split(',')
                if not first_line[0].replace('.', '').replace('-', '').isdigit():
                    start_index = 1

                rows = []
                for i, line in enumerate(lines[start_index:], 1):
                    values = [v.strip() for v in line.split(',')]
                    if len(values) < 2:
                        continue

                    distance = float(values[0]) if values[0] else 0
                    depth = float(values[1]) if values[1] else 0

                    # 측정법 자동 결정
                    if depth == 0:
                        method = 'LEW' if i == 1 else 'REW'
                    elif depth < 0.6:
                        method = '1'
                    elif depth < 1.0:
                        method = '2'
                    else:
                        method = '3'

                    row = {
                        'id': i,
                        'distance': distance,
                        'depth': depth,
                        'method': method,
                        'index_value': float(values[2]) if len(values) > 2 and values[2] else 1.0,
                        'n_02d': float(values[3]) if len(values) > 3 and values[3] else None,
                        't_02d': float(values[4]) if len(values) > 4 and values[4] else None,
                        'n_06d': float(values[5]) if len(values) > 5 and values[5] else None,
                        't_06d': float(values[6]) if len(values) > 6 and values[6] else None,
                        'n_08d': float(values[7]) if len(values) > 7 and values[7] else None,
                        't_08d': float(values[8]) if len(values) > 8 and values[8] else None,
                    }
                    rows.append(row)

                if not rows:
                    errors.append({'filename': filename, 'error': '유효한 데이터가 없습니다.'})
                    continue

                # 중복 체크: 같은 관측소 + 날짜가 이미 존재하면 skip
                if station_name and measurement_date:
                    existing = MeasurementSession.objects.filter(
                        station_name=station_name,
                        measurement_date=measurement_date
                    ).first()
                    if existing:
                        errors.append({
                            'filename': filename,
                            'error': f'중복 데이터 (이미 존재: {station_name}, {measurement_date})',
                            'duplicate': True,
                            'existing_session_id': existing.pk
                        })
                        continue

                # 유량 계산
                result_data = calculate_discharge(rows, calibration)

                # MeasurementSession 생성
                session = MeasurementSession.objects.create(
                    user=user,
                    session_key=session_key,
                    station_name=station_name,
                    measurement_date=measurement_date,
                    rows_data=rows,
                    calibration_data=calibration,
                    setup_data={
                        'river_name': river_name,
                        'location': location,
                        'source_file': filename,
                        'final_discharge': result_data.get('discharge'),
                        'final_uncertainty': result_data.get('uncertainty'),
                        'final_avg_velocity': result_data.get('avg_velocity'),
                    },
                    estimated_discharge=result_data.get('discharge'),
                    total_width=result_data.get('width'),
                    max_depth=result_data.get('avg_depth'),
                    total_area=result_data.get('area'),
                )

                # 분석 결과 계산 및 저장
                session.calculate_analysis_results()
                session.save()

                results.append({
                    'filename': filename,
                    'session_id': session.pk,
                    'station_name': station_name,
                    'measurement_date': measurement_date.strftime('%Y-%m-%d') if measurement_date else None,
                    'discharge': result_data.get('discharge'),
                    'area': result_data.get('area'),
                    'width': result_data.get('width'),
                    'uncertainty': result_data.get('uncertainty'),
                })

            except Exception as e:
                errors.append({'filename': filename, 'error': str(e)})

        return JsonResponse({
            'success': True,
            'imported': len(results),
            'failed': len(errors),
            'results': results,
            'errors': errors,
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
